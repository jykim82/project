from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 공통 스타일 ──
HEADER_FONT = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
SUB_HEADER_FILL = PatternFill('solid', fgColor='4472C4')
CATEGORY_FILL = PatternFill('solid', fgColor='D6E4F0')
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14)
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=12)
NORMAL_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', bold=True, size=10)
TOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')
SUBTOTAL_FILL = PatternFill('solid', fgColor='E2EFDA')
thin = Side(style='thin', color='B4C6E7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
NUM_FMT = '#,##0'
MONEY_FMT = '#,##0'

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_row(ws, row, cols, fill=None, font=None, aligns=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or NORMAL_FONT
        cell.border = BORDER
        if fill: cell.fill = fill
        if aligns and c <= len(aligns):
            cell.alignment = aligns[c-1]
        else:
            cell.alignment = CENTER

def add_title(ws, title, row=1, col=1, merge_to=8):
    ws.cell(row=row, column=col, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)

# ═══════════════════════════════════════
# Sheet 1: 대가산정 총괄표
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = '대가산정 총괄표'
ws1.sheet_properties.tabColor = '2F5496'

add_title(ws1, 'SLM 스마트 관망관리 시스템 — 대가산정 총괄표')
ws1.cell(row=2, column=1, value='작성일: 2026-03-14').font = Font(name='맑은 고딕', size=10, color='666666')

headers = ['구분', '항목', '세부 내용', '규모', '단위', '단가(원)', '금액(원)', '비고']
r = 4
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8)

data = [
    # (구분, 항목, 세부내용, 규모, 단위, 단가, 금액, 비고)
    ('1. 소프트웨어 개발', '프론트엔드 개발', 'Next.js 14 + TypeScript + Tailwind', 34, '화면', 2500000, None, 'App Router, shadcn/ui'),
    ('', '프론트엔드 컴포넌트', 'React 컴포넌트 + 커스텀 훅', 163, '개', 800000, None, 'UI/CRUD/차트/캔버스'),
    ('', '차트·시각화 개발', 'ECharts + SVG 커스텀 차트', 14, '종', 4000000, None, '트렌드/토폴로지/계통도/포트'),
    ('', 'API 클라이언트', '프론트엔드 API 연동 모듈', 27, '개', 500000, None, 'fetch 래퍼 + 타입'),
    ('', '상태 관리', 'Zustand 스토어', 17, '개', 600000, None, '전역 상태 관리'),
    ('', 'Backend API 개발', 'FastAPI 엔드포인트', 104, '개', 1500000, None, 'CRUD + 분석 + SSE'),
    ('', 'AI 인텐트 분류', '3단계 분류(키워드→벡터→SLM)', 68, '인텐트', 2000000, None, '501질문, snowflake-arctic'),
    ('', '이상감지 엔진', '4계층 탐지 (인과+홀딩+Z-Score+물수지)', 4, '계층', 15000000, None, 'Rule+ML+통계+물리'),
    ('', '인과관계 엔진', '물리법칙 기반 Rule 엔진', 1, '식', 20000000, None, '6규칙+교차검증+전파추적'),
    ('', '실시간 모니터링', '용수 흐름 계통도 + 타임라인', 1, '식', 18000000, None, 'SVG+파티클+미니맵+LOD'),
    ('', '캔버스 에디터', 'React Flow 기반 시설 토폴로지', 1, '식', 12000000, None, 'Undo/Redo+자동배치+내보내기'),
    ('2. DB 설계·구축', '테이블 설계', 'PostgreSQL + TimescaleDB', 50, '테이블', 800000, None, '정규화 + 하이퍼테이블'),
    ('', '뷰·집계 설계', 'View + Continuous Aggregate', 7, '개', 1200000, None, '통합뷰 + 5분 집계'),
    ('', '데이터 마이그레이션', '운영DB → 로컬DB 동기화', 1, '식', 5000000, None, 'db_sync.py 증분 동기화'),
    ('', '함수·프로시저', 'PostgreSQL 사용자 함수', 13, '개', 1500000, None, '야간최소유량/결측/통계'),
    ('3. 인프라·환경', 'Docker 환경', 'TimescaleDB + Node-RED', 2, '컨테이너', 3000000, None, 'docker-compose'),
    ('', 'HTTPS 인증서', 'mkcert 로컬 SSL', 1, '식', 1000000, None, '브라우저↔Next.js HTTPS'),
    ('', 'Cloudflare Tunnel', '외부 접속 터널', 1, '식', 1000000, None, 'Quick Tunnel'),
    ('4. 테스트·검증', '인텐트 전수 테스트', '68개 인텐트 E2E 검증', 68, '건', 300000, None, 'test_all_intents.py'),
    ('', '성능 최적화', '쿼리 최적화 + 캐시', 8, '건', 3000000, None, '청크 직접쿼리, 배경캐시'),
    ('', 'Playwright E2E', '브라우저 자동화 테스트', 1, '식', 5000000, None, 'UI 검증'),
]

for i, d in enumerate(data):
    r = 5 + i
    for c, v in enumerate(d, 1):
        ws1.cell(row=r, column=c, value=v)
    aligns = [LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, RIGHT, CENTER, RIGHT, RIGHT, LEFT_WRAP]
    fill = CATEGORY_FILL if d[0] else None
    style_row(ws1, r, 8, fill=fill, font=BOLD_FONT if d[0] else NORMAL_FONT, aligns=aligns)
    # 금액 = 규모 × 단가 (열 7 = 열4 × 열6)
    ws1.cell(row=r, column=7, value=f'=D{r}*F{r}')
    ws1.cell(row=r, column=7).number_format = MONEY_FMT
    ws1.cell(row=r, column=6).number_format = MONEY_FMT
    ws1.cell(row=r, column=4).number_format = NUM_FMT

# 소계 행들
r_subtotal = 5 + len(data)
subtotals = [
    ('소프트웨어 개발 소계', 5, 15),
    ('DB 설계·구축 소계', 16, 19),
    ('인프라·환경 소계', 20, 21),
    ('테스트·검증 소계', 22, 24),
]
for label, s, e in subtotals:
    ws1.cell(row=r_subtotal, column=1, value=label)
    ws1.merge_cells(start_row=r_subtotal, start_column=1, end_row=r_subtotal, end_column=5)
    ws1.cell(row=r_subtotal, column=7, value=f'=SUM(G{s}:G{e})')
    ws1.cell(row=r_subtotal, column=7).number_format = MONEY_FMT
    style_row(ws1, r_subtotal, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT)
    ws1.cell(row=r_subtotal, column=1).alignment = LEFT_WRAP
    ws1.cell(row=r_subtotal, column=7).alignment = RIGHT
    r_subtotal += 1

# 합계
r_total = r_subtotal
ws1.cell(row=r_total, column=1, value='합 계')
ws1.merge_cells(start_row=r_total, start_column=1, end_row=r_total, end_column=5)
sub_rows = [r_total-4, r_total-3, r_total-2, r_total-1]
ws1.cell(row=r_total, column=7, value=f'=G{sub_rows[0]}+G{sub_rows[1]}+G{sub_rows[2]}+G{sub_rows[3]}')
ws1.cell(row=r_total, column=7).number_format = MONEY_FMT
style_row(ws1, r_total, 8, fill=TOTAL_FILL, font=Font(name='맑은 고딕', bold=True, size=11))
ws1.cell(row=r_total, column=1).alignment = LEFT_WRAP
ws1.cell(row=r_total, column=7).alignment = RIGHT

# 부가세, 총합계
r_vat = r_total + 1
ws1.cell(row=r_vat, column=1, value='부가가치세 (10%)')
ws1.merge_cells(start_row=r_vat, start_column=1, end_row=r_vat, end_column=5)
ws1.cell(row=r_vat, column=7, value=f'=G{r_total}*0.1')
ws1.cell(row=r_vat, column=7).number_format = MONEY_FMT
style_row(ws1, r_vat, 8, font=BOLD_FONT)
ws1.cell(row=r_vat, column=1).alignment = LEFT_WRAP
ws1.cell(row=r_vat, column=7).alignment = RIGHT

r_grand = r_vat + 1
ws1.cell(row=r_grand, column=1, value='총 합 계 (VAT 포함)')
ws1.merge_cells(start_row=r_grand, start_column=1, end_row=r_grand, end_column=5)
ws1.cell(row=r_grand, column=7, value=f'=G{r_total}+G{r_vat}')
ws1.cell(row=r_grand, column=7).number_format = MONEY_FMT
style_row(ws1, r_grand, 8, fill=PatternFill('solid', fgColor='FFD966'), font=Font(name='맑은 고딕', bold=True, size=12))
ws1.cell(row=r_grand, column=1).alignment = LEFT_WRAP
ws1.cell(row=r_grand, column=7).alignment = RIGHT

for c, w in enumerate([18, 20, 35, 8, 8, 14, 16, 25], 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════
# Sheet 2: 화면 목록
# ═══════════════════════════════════════
ws2 = wb.create_sheet('화면 목록')
ws2.sheet_properties.tabColor = '4472C4'
add_title(ws2, '화면(페이지) 목록', merge_to=7)

headers2 = ['No', '메뉴 그룹', '화면명', '경로', '주요 기능', '난이도', '비고']
r = 3
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 7)

screens = [
    ('랜딩', '랜딩 페이지', '/', '시스템 소개, 로그인 연결', '하', ''),
    ('인증', '로그인', '/login', 'NextAuth 인증, JWT', '중', ''),
    ('대시보드', '종합 현황판', '/dashboard', 'KPI 6종, 이상시설 TOP, 유량불균형, 설비장애, 데이터품질, 최근경보, 클릭→팝업 분석', '상', '캐시 집계'),
    ('AI 채팅', 'AI 어시스턴트', '/chat', 'SSE 스트리밍, 68인텐트, 시각화 렌더링, 히스토리, 자동질문', '최상', '핵심 기능'),
    ('모니터링', '배수지 모니터링', '/monitoring/reservoir', '트렌드 차트, HH/LL 가이드선, 재생, 네비게이션', '상', ''),
    ('', '가압장 모니터링', '/monitoring/booster', '트렌드 차트, 알람 마크라인', '상', ''),
    ('', '감압시설 모니터링', '/monitoring/pressure', '트렌드 차트', '중', ''),
    ('', '블록 모니터링', '/monitoring/block', '트렌드 차트', '중', ''),
    ('', '용수 흐름 모니터링', '/monitoring/flow', 'SVG 계통도, 실시간 유량, 파티클, 미니맵, 검색, LOD, 타임라인, 트렌드 패널, KPI 7종', '최상', '핵심 기능'),
    ('트렌드', '트렌드 분석', '/trend', '다중 태그 비교, 태그 브라우저, 이동평균', '상', ''),
    ('네트워크', '네트워크 토폴로지', '/network', 'Force/계층형 듀얼 뷰, SNMP 포트, 노드 상세', '상', ''),
    ('알람', '알람 알림', '/alarm', '자동 스캔, 실시간 알림', '중', ''),
    ('위기대응', '경보관리 (현황/이력)', '/crisis/alarm-dashboard', '2탭 (도넛+카테고리+테이블 / 필터+확인)', '상', ''),
    ('', '경보분석', '/crisis/alarm-analysis', '경보분석 상세, 진단 메시지, iframe', '상', ''),
    ('', '경보이력 (리다이렉트)', '/crisis/alarm-history', '→ alarm-dashboard?tab=history', '하', ''),
    ('', '현장조치', '/crisis/task-management', '조치 관리', '중', ''),
    ('구축', '태그 마스터', '/setup/tags', '2,698건, 5종 필터, 페이징, CSV 업/다운', '상', ''),
    ('', '설비 관리', '/setup/equipments', '290건, 3종 필터, 자동ID, CSV 업/다운', '상', ''),
    ('', '배수지 정보', '/setup/reservoir', '3탭 폼(기본/구역/운영), 설비메타 16항목, CSV', '상', ''),
    ('', '가압장 정보', '/setup/booster', '2탭 폼(기본/운영), 설비메타 26항목, CSV', '상', ''),
    ('', '감압시설 정보', '/setup/pressure', '2탭 폼, 설비메타 5항목, CSV', '중', ''),
    ('', '블록 정보', '/setup/block', '2탭 폼, 블록레벨 필터, CSV', '중', ''),
    ('', '용수 흐름 계통도', '/setup/flow-map', 'SVG 계통도, CRUD, CSV 입출력', '상', ''),
    ('', '네트워크 관리', '/setup/networks', '장비탭+연결탭, 토폴로지 미리보기, 미니 플로우', '상', ''),
    ('', '트렌드 설정', '/setup/trends', '모니터링 카탈로그 CRUD', '중', ''),
    ('', '캔버스 에디터', '/setup/canvas', 'React Flow, 78노드, Undo/Redo, 자동배치, PNG/SVG 내보내기, 인과 탭', '최상', 'Phase 1-4'),
    ('', '인과 규칙', '/setup/causal-rules', '5탭 템플릿, 시설 커버리지, 태그 매핑 현황', '상', ''),
    ('', '컬럼 잠금', '/setup/field-locks', '컬럼별 접근 제어', '중', ''),
    ('관리자', '사용자 관리', '/admin/users', 'CRUD, 역할 배정', '중', ''),
    ('', '메뉴 관리', '/admin/menus', '트리 구조 CRUD', '중', ''),
    ('', 'FAQ 관리', '/admin/faq', 'CRUD', '중', ''),
    ('', '프롬프트 관리', '/admin/prompts', '인텐트별 프롬프트 편집', '중', ''),
    ('', '시설 파일 관리', '/admin/facility-files', '이미지 업로드, 파일 관리', '중', ''),
    ('', '사이트 설정', '/admin/site-settings', '시스템 설정', '하', ''),
]

for i, s in enumerate(screens):
    r = 4 + i
    ws2.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(s, 2):
        ws2.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, CENTER, LEFT_WRAP]
    style_row(ws2, r, 7, fill=CATEGORY_FILL if s[0] else None, font=BOLD_FONT if s[0] else NORMAL_FONT, aligns=aligns)

for c, w in enumerate([5, 12, 20, 30, 45, 8, 15], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════
# Sheet 3: API 목록
# ═══════════════════════════════════════
ws3 = wb.create_sheet('API 목록')
ws3.sheet_properties.tabColor = '70AD47'
add_title(ws3, 'API 엔드포인트 목록 (FastAPI Backend)', merge_to=7)

headers3 = ['No', '카테고리', 'Method', 'URL', '설명', '난이도', '비고']
r = 3
for c, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 7)

apis = [
    ('채팅/AI', 'POST', '/ask', '동기 채팅 쿼리 (인텐트 분류→SQL→응답)', '최상', '68인텐트'),
    ('', 'POST', '/ask/stream', 'SSE 스트리밍 채팅 (토큰→answer→visual→recommend→done)', '최상', 'SSE'),
    ('대시보드', 'GET', '/dashboard/overview', '종합 현황판 (캐시 집계+경보+교차검증+물수지+품질+장애)', '상', '캐시 집계'),
    ('모니터링', 'GET', '/monitoring/dashboard', '모니터링 대시보드 요약', '중', ''),
    ('', 'GET', '/monitoring/alarm-notifications', '알람 알림 (진행중 건수+최근 5건)', '중', ''),
    ('', 'GET', '/monitoring/catalogs/sites', '모니터링 사이트 목록', '중', ''),
    ('', 'GET', '/monitoring/catalogs', '모니터링 카탈로그 조회', '중', ''),
    ('', 'GET', '/monitoring/catalogs/reference', '카탈로그 참조 데이터', '중', ''),
    ('', 'POST', '/monitoring/catalogs', '모니터링 카탈로그 생성', '중', ''),
    ('', 'PUT', '/monitoring/catalogs/{id}', '모니터링 카탈로그 수정', '중', ''),
    ('', 'DELETE', '/monitoring/catalogs/{id}', '모니터링 카탈로그 삭제', '중', ''),
    ('태그', 'GET', '/tags', '태그 마스터 조회 (5종 필터+페이징)', '중', '2,698건'),
    ('', 'GET', '/tags/filters', '태그 필터 옵션', '하', ''),
    ('', 'GET', '/tags/groups', '태그 그룹 통계 (21그룹)', '중', ''),
    ('', 'POST', '/tags/import/csv', '태그 CSV 일괄 등록', '중', ''),
    ('트렌드', 'POST', '/trend/data', '시계열 데이터 조회 (time_bucket, 최대 15태그)', '상', '청크 최적화'),
    ('이상감지', 'GET', '/anomaly/profiles', '현장 프로파일 조회 (A/B/C/D 그룹)', '중', '디버깅'),
    ('인과관계', 'GET', '/causal/rules', '인과 체인 템플릿+시설 매핑 현황', '상', ''),
    ('', 'GET', '/causal/verify', '인과 체인 검증 (디버그)', '중', ''),
    ('', 'GET', '/causal/chain/{sn}/{ft}', '인과 체인 상세 조회', '중', ''),
    ('', 'PUT', '/causal/chain/{sn}/{ft}', '인과 체인 오버라이드 저장', '상', ''),
    ('', 'DELETE', '/causal/chain/{sn}/{ft}', '인과 체인 오버라이드 삭제', '하', ''),
    ('', 'POST', '/causal/estimate-lag', '교차상관 시간 지연 추정', '상', 'numpy'),
    ('용수흐름', 'GET', '/flow-map', '전체 계통도 (토폴로지+메타)', '중', '95엣지'),
    ('', 'GET', '/flow-map/realtime', '실시간 유량/수위/압력+교차검증+물수지+장애+공급시간', '최상', '핵심'),
    ('', 'GET', '/flow-map/node-alarms', '시설별 최근 알람', '중', ''),
    ('', 'GET', '/flow-map/roots', '최상위 노드 (상류)', '하', ''),
    ('', 'GET', '/flow-map/downstream', '하류 계단식 조회 (재귀 CTE)', '중', ''),
    ('', 'POST', '/flow-map', '계통도 연결 추가', '중', ''),
    ('', 'DELETE', '/flow-map', '계통도 연결 삭제', '하', ''),
    ('', 'GET', '/flow-map/export/csv', '계통도 CSV 내보내기', '하', ''),
    ('', 'POST', '/flow-map/import/csv', '계통도 CSV 가져오기', '중', ''),
    ('설비', 'GET', '/equipments', '설비 목록 (3종 필터+페이징)', '중', '290건'),
    ('', 'GET', '/equipments/filters', '설비 필터 옵션', '하', ''),
    ('', 'GET', '/equipments/next-id', '설비 자동 ID 생성', '중', ''),
    ('', 'POST', '/equipments', '설비 등록', '중', ''),
    ('', 'PUT', '/equipments/{id}', '설비 수정', '중', ''),
    ('', 'DELETE', '/equipments/{id}', '설비 삭제 (CASCADE 체크)', '중', 'dry_run'),
    ('', 'GET', '/equipments/auto-map', '설비↔태그 자동 매핑', '상', '3,375건'),
    ('', 'POST', '/equipments/import/csv', '설비 CSV 일괄 등록', '중', ''),
    ('배수지', 'GET', '/reservoirs', '배수지 목록', '중', ''),
    ('', 'GET', '/reservoirs/{sn}', '배수지 상세', '중', ''),
    ('', 'POST', '/reservoirs', '배수지 등록 (info+status)', '상', '27컬럼'),
    ('', 'PUT', '/reservoirs/{sn}', '배수지 수정', '상', ''),
    ('', 'DELETE', '/reservoirs/{sn}', '배수지 삭제', '중', 'dry_run'),
    ('', 'POST', '/reservoirs/import/csv', '배수지 CSV 일괄 등록', '중', '27컬럼'),
    ('가압장', 'GET', '/boosters', '가압장 목록', '중', ''),
    ('', 'GET', '/boosters/{sn}', '가압장 상세', '중', ''),
    ('', 'POST', '/boosters', '가압장 등록', '상', '18컬럼'),
    ('', 'PUT', '/boosters/{sn}', '가압장 수정', '상', ''),
    ('', 'DELETE', '/boosters/{sn}', '가압장 삭제', '중', ''),
    ('', 'POST', '/boosters/import/csv', '가압장 CSV 일괄 등록', '중', ''),
    ('감압시설', 'GET', '/pressure-reducing', '감압시설 목록', '중', ''),
    ('', 'GET', '/pressure-reducing/{sn}', '감압시설 상세', '중', ''),
    ('', 'POST', '/pressure-reducing', '감압시설 등록', '중', ''),
    ('', 'PUT', '/pressure-reducing/{sn}', '감압시설 수정', '중', ''),
    ('', 'DELETE', '/pressure-reducing/{sn}', '감압시설 삭제', '중', ''),
    ('', 'POST', '/pressure-reducing/import/csv', '감압시설 CSV 일괄 등록', '중', ''),
    ('블록', 'GET', '/blocks', '블록 목록 (블록레벨 필터)', '중', ''),
    ('', 'GET', '/blocks/{sn}', '블록 상세', '중', ''),
    ('', 'POST', '/blocks', '블록 등록', '중', ''),
    ('', 'PUT', '/blocks/{sn}', '블록 수정', '중', ''),
    ('', 'DELETE', '/blocks/{sn}', '블록 삭제', '중', ''),
    ('', 'POST', '/blocks/import/csv', '블록 CSV 일괄 등록', '중', ''),
    ('네트워크', 'GET', '/network/devices', '네트워크 장비 + 최신 상태', '중', 'IP 기반'),
    ('', 'GET', '/network/topology', '토폴로지 그래프 (노드+엣지)', '중', ''),
    ('', 'GET', '/network/status/summary', '네트워크 상태 요약', '중', ''),
    ('', 'GET', '/network/infos', '네트워크 장비 상세 (CRUD)', '중', ''),
    ('', 'GET', '/network/infos/filters', '네트워크 필터 옵션', '하', ''),
    ('', 'POST', '/network/infos', '네트워크 장비 등록', '중', ''),
    ('', 'PUT', '/network/infos/{id}', '네트워크 장비 수정', '중', ''),
    ('', 'DELETE', '/network/infos/{id}', '네트워크 장비 삭제 (CASCADE)', '중', ''),
    ('', 'GET', '/network/links', '네트워크 연결 목록', '중', ''),
    ('', 'GET', '/network/links/protocols', '프로토콜 마스터', '하', ''),
    ('', 'POST', '/network/links', '네트워크 연결 추가', '중', ''),
    ('', 'PUT', '/network/links/{s}/{t}', '네트워크 연결 수정', '중', ''),
    ('', 'DELETE', '/network/links/{s}/{t}', '네트워크 연결 삭제', '하', ''),
    ('', 'GET', '/network/links/equipment-search', '장비 자동완성', '중', ''),
    ('', 'POST', '/network/devices/import/csv', '네트워크 장비 CSV 등록', '중', ''),
    ('', 'POST', '/network/links/import/csv', '네트워크 연결 CSV 등록', '중', ''),
    ('SNMP', 'GET', '/network/snmp/{id}/ports', '스위치 포트 상태', '중', '24포트'),
    ('', 'GET', '/network/snmp/{id}/system', '스위치 시스템 정보', '중', ''),
    ('', 'GET', '/network/snmp/summary', '전체 스위치 포트 요약', '중', '13대'),
    ('위기대응', 'GET', '/crisis/alarm-reports', '경보 리포트 (날짜 필터+페이징)', '중', ''),
    ('', 'GET', '/crisis/alarm-analysis', '경보분석 데이터', '상', '진단 메시지'),
    ('', 'GET', '/crisis/alarm-analysis/detail', '경보분석 단건 상세', '중', ''),
    ('', 'GET', '/crisis/alarm-dashboard', '경보 대시보드 요약', '중', ''),
    ('', 'PUT', '/crisis/alarm-reports/confirm', '경보 확인 처리', '하', ''),
    ('캔버스', 'GET', '/canvas/layout', '캔버스 레이아웃 (노드+엣지+설비+태그)', '상', '78노드'),
    ('', 'PUT', '/canvas/layout', '캔버스 레이아웃 저장 (UPSERT+엣지 diff)', '상', ''),
    ('', 'GET', '/canvas/node-detail/{sn}/{ft}', '노드 상세 (설비+카탈로그)', '중', ''),
    ('', 'GET', '/canvas/equipment-tags/{sn}/{ft}', '설비별 태그 목록', '중', ''),
    ('', 'POST', '/canvas/equipment-tag-link', '설비↔태그 연결', '중', ''),
    ('', 'DELETE', '/canvas/equipment-tag-link', '설비↔태그 해제', '하', ''),
    ('관리자', 'POST', '/admin/facility-files/upload', '시설 이미지 업로드 (multipart)', '상', ''),
    ('', 'GET', '/admin/facility-files', '시설 파일 목록', '중', ''),
    ('', 'GET', '/admin/facilities-summary', '시설 + 파일 등록 현황', '중', ''),
    ('', 'DELETE', '/admin/facility-files/{id}', '시설 파일 삭제', '하', ''),
    ('', 'GET', '/admin/site-settings', '사이트 설정 조회', '하', ''),
    ('', 'PUT', '/admin/site-settings', '사이트 설정 수정', '하', ''),
    ('유틸리티', 'GET', '/autocomplete/candidates', '자동완성 (현장명/시설유형/데이터항목)', '중', ''),
    ('', 'GET', '/csv/{filename}', 'CSV 파일 다운로드', '하', ''),
    ('', 'GET', '/health', '서버 헬스체크 (Ollama+DB)', '하', ''),
    ('', 'GET', '/models', 'Ollama 모델 목록', '하', ''),
    ('', 'POST', '/models/select', 'Ollama 모델 변경', '하', ''),
]

for i, a in enumerate(apis):
    r = 4 + i
    ws3.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(a, 2):
        ws3.cell(row=r, column=c, value=v)
    method_col = ws3.cell(row=r, column=3)
    if a[1] == 'POST': method_col.fill = PatternFill('solid', fgColor='E2F0D9')
    elif a[1] == 'PUT': method_col.fill = PatternFill('solid', fgColor='FFF2CC')
    elif a[1] == 'DELETE': method_col.fill = PatternFill('solid', fgColor='FCE4EC')
    aligns = [CENTER, LEFT_WRAP, CENTER, LEFT_WRAP, LEFT_WRAP, CENTER, LEFT_WRAP]
    style_row(ws3, r, 7, fill=CATEGORY_FILL if a[0] else None, font=BOLD_FONT if a[0] else NORMAL_FONT, aligns=aligns)

r_api_total = 4 + len(apis)
ws3.cell(row=r_api_total, column=1, value=f'합계: {len(apis)}개 API')
ws3.merge_cells(start_row=r_api_total, start_column=1, end_row=r_api_total, end_column=7)
style_row(ws3, r_api_total, 7, fill=TOTAL_FILL, font=BOLD_FONT)

for c, w in enumerate([5, 10, 8, 35, 42, 8, 15], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════
# Sheet 4: DB 테이블 목록
# ═══════════════════════════════════════
ws4 = wb.create_sheet('DB 테이블 목록')
ws4.sheet_properties.tabColor = 'ED7D31'
add_title(ws4, 'DB 테이블 설계 목록 (PostgreSQL + TimescaleDB)', merge_to=7)

headers4 = ['No', '카테고리', '테이블/뷰명', '유형', '설명', '주요 컬럼', '비고']
r = 3
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 7)

tables = [
    ('인증/사용자', 'tb_auth', '테이블', '권한 그룹 (ADMIN, USER, VIEWER)', 'auth_idn, auth_name, auth_level', ''),
    ('', 'tb_user', '테이블', '사용자 계정 (bcrypt + AES 마이그레이션)', 'user_id, pw, pw_migrated, auth_idn, region', ''),
    ('', 'tb_user_session', '테이블', '로그인 세션 관리', 'session_id, user_id, expires_at', ''),
    ('', 'tb_access_log', '테이블', 'API 접근 감사 로그', 'log_id, user_id, api_path, timestamp', ''),
    ('메뉴/권한', 'tb_menu', '테이블', '메뉴 트리 (자기참조 FK)', 'menu_idn, pmenu_idn, menu_name, app_path', ''),
    ('', 'tb_auth_menu', '테이블', '권한↔메뉴 매핑', 'auth_idn, menu_idn, read_yn, write_yn', ''),
    ('', 'tb_menu_api', '테이블', '메뉴↔API 매핑', 'menu_idn, api_path, method', ''),
    ('공통코드', 'tb_grp_code', '테이블', '공통 그룹 코드 마스터', 'grp_code, grp_name', ''),
    ('', 'tb_comm_code', '테이블', '공통 상세 코드', 'grp_code, comm_code, comm_name', ''),
    ('AI 채팅', 'tb_ai_chat_ask_group', '테이블', '채팅 세션 그룹', 'group_id, user_id, title, created_at', ''),
    ('', 'tb_ai_chat_ask', '테이블', '사용자 질문 메시지', 'ask_id, group_id, question, ask_at', ''),
    ('', 'tb_ai_chat_bot', '테이블', '봇 응답 (visual_data JSONB)', 'bot_id, ask_id, answer, visual_data, bot_at', '차트 재렌더링'),
    ('', 'tb_ai_chat_ask_image', '테이블', '질문 첨부 이미지', 'image_id, ask_id, file_path', ''),
    ('', 'tb_ai_chat_bot_image', '테이블', '응답 첨부 이미지', 'image_id, bot_id, file_path', ''),
    ('', 'tb_ai_chat_faq', '테이블', 'FAQ/추천 질문', 'faq_id, question, category', ''),
    ('시설정보', 'tb_service_reservoir_info', '테이블', '배수지 기본정보 (general_overview JSONB)', 'sitename, region, facilitytype, install_location, capacity_m3', ''),
    ('', 'tb_service_reservoir_status', '테이블', '배수지 운영 상태', 'sitename, hh, ll, target_level, daily_supply, meta', '설비메타 16항목'),
    ('', 'tb_service_booster_station_info', '테이블', '가압장 기본정보', 'sitename, region, facilitytype, install_location', ''),
    ('', 'tb_service_booster_station_status', '테이블', '가압장 운영 상태', 'sitename, hh, ll, target_pressure, meta', '설비메타 26항목'),
    ('', 'tb_pressure_reducing_facility_info', '테이블', '감압시설 기본정보', 'sitename, region, facilitytype', ''),
    ('', 'tb_pressure_reducing_facility_status', '테이블', '감압시설 운영 상태', 'sitename, hh, ll, meta', '설비메타 5항목'),
    ('', 'tb_block_info', '테이블', '블록 기본정보 (소블록/소소블록)', 'sitename, block_level, region', ''),
    ('', 'tb_block_status', '테이블', '블록 운영 상태', 'sitename, target_flow, meta', '설비메타 5항목'),
    ('설비/장비', 'tb_equipment_info', '테이블', '설비 마스터 (센서, 제어기, 네트워크)', 'equipment_id, sitename, equipment_type, ip_address', '290건'),
    ('', 'tb_equipment_status', '테이블', '설비 운영 상태 (is_alive, RTT)', 'equipment_id, is_alive, rtt_ms, status_code', ''),
    ('', 'tb_equipment_alarm_report', '테이블', '설비 경보 리포트 (해결 추적)', 'alarm_id, equipment_id, alarm_type, start_time, resolved', ''),
    ('', 'tb_equipment_tag_map', '테이블', '설비↔태그 매핑', 'equipment_id, tagsn, mapped_by', '3,375건 자동'),
    ('네트워크', 'tb_network_link', '테이블', '네트워크 장비 연결 (방향성)', 'source_id, target_id, protocol', ''),
    ('', 'tb_facility_flow_map', '테이블', '용수 계통도 연결', 'source_sitename, target_sitename, source_ft, target_ft', '95엣지'),
    ('태그/시계열', 'tb_tag_info', '테이블', '센서/태그 마스터', 'tagsn, sitename, facilitytype, tag_type, datainfo, unit', '2,698건'),
    ('', 'tb_tag_raw_data', '하이퍼테이블', '시계열 센서 데이터 (TimescaleDB)', 'tagsn, timestamp, value, quality', '2.3M+ 행'),
    ('', 'tb_tag_data_group', '테이블', '태그 데이터 그룹 분류 (21그룹)', 'group_code, parent_code, group_name, keywords', '계층 구조'),
    ('', 'tb_tag_group_map', '테이블', '태그↔그룹 매핑', 'tagsn, group_code, matched_keyword', '93% 자동'),
    ('', 'tb_trend_catalog', '테이블', '트렌드 차트 카탈로그', 'catalog_id, name, items, facilitytype', ''),
    ('', 'tb_alarm_log', '테이블', '센서 알람 이력 (HH/LL/H/L)', 'tagsn, alarm_type, start_time, end_time, resolved', ''),
    ('모니터링', 'tb_monitoring_catalog', '테이블', '모니터링 전용 카탈로그', 'catalog_id, catalog_name, display_order, items', '분리 운영'),
    ('', 'tb_admin_site_settings', '테이블', '사이트 관리 설정', 'setting_key, setting_value', ''),
    ('', 'tb_site_anomaly_profile', '테이블', '현장 이상감지 프로파일', 'sitename, group(A/B/C/D), p95, p05, alarm_freq', '일 1회 갱신'),
    ('캔버스/인과', 'tb_canvas_node_position', '테이블', '캔버스 노드 위치', 'sitename, facilitytype, x, y', ''),
    ('', 'tb_causal_chain_override', '테이블', '인과 규칙 오버라이드', 'sitename, facilitytype, zone, chain, cross_facility', 'JSONB'),
    ('', 'tb_snmp_port_status', '테이블', 'SNMP 스위치 포트 상태', 'equipment_id, port_index, status, speed, traffic', '13대×24포트'),
    ('관리자/파일', 'tb_prompt_template', '테이블', 'AI 프롬프트 템플릿', 'intent, template, column_defs', ''),
    ('', 'tb_prompt_column', '테이블', '프롬프트 컬럼 정의', 'intent, column_name, column_type', ''),
    ('', 'tb_file_storage', '테이블', '파일 저장 메타데이터', 'file_id, file_path, mime_type, category', ''),
    ('', 'tb_file_history', '테이블', '파일 버전 이력', 'history_id, file_id, version, uploaded_at', ''),
    ('', 'tb_facility_file', '테이블', '시설↔파일 매핑', 'facility_file_id, sitename, file_id, file_type', ''),
    ('', 'tb_field_lock', '테이블', '컬럼 접근 제어', 'table_name, column_name, locked_by', ''),
    ('뷰/집계', 'v_reservoir_info_status', '뷰', '배수지 info + status + 파일 URL 통합', '', ''),
    ('', 'v_booster_station_info_status', '뷰', '가압장 info + status + 파일 URL 통합', '', ''),
    ('', 'v_pressure_reducing_facility_info_status', '뷰', '감압시설 info + status + 파일 URL 통합', '', ''),
    ('', 'v_block_info_status', '뷰', '블록 info + status + 파일 URL 통합', '', ''),
    ('', 'v_ongoing_alarm', '뷰', '진행 중(미해결) 알람', '', ''),
    ('', 'cagg_5min_raw_stats_ai', '연속집계', 'TimescaleDB 5분 통계 (avg/min/max)', '', 'Continuous Aggregate'),
    ('', 'mv_tag_daily_status', '물리화뷰', '일별 태그 통계 (avg/min/max)', '', 'Materialized View'),
]

for i, t in enumerate(tables):
    r = 4 + i
    ws4.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(t, 2):
        ws4.cell(row=r, column=c, value=v)
    type_cell = ws4.cell(row=r, column=4)
    if t[2] == '하이퍼테이블': type_cell.fill = PatternFill('solid', fgColor='E8D5F5')
    elif t[2] in ('뷰', '연속집계', '물리화뷰'): type_cell.fill = PatternFill('solid', fgColor='D5E8F5')
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, CENTER, LEFT_WRAP, LEFT_WRAP, LEFT_WRAP]
    style_row(ws4, r, 7, fill=CATEGORY_FILL if t[0] else None, font=BOLD_FONT if t[0] else NORMAL_FONT, aligns=aligns)

r_db_total = 4 + len(tables)
ws4.cell(row=r_db_total, column=1, value=f'합계: 테이블 {sum(1 for t in tables if t[2]=="테이블")}개 + 뷰/집계 {sum(1 for t in tables if t[2]!="테이블")}개 = {len(tables)}개')
ws4.merge_cells(start_row=r_db_total, start_column=1, end_row=r_db_total, end_column=7)
style_row(ws4, r_db_total, 7, fill=TOTAL_FILL, font=BOLD_FONT)

for c, w in enumerate([5, 10, 38, 10, 40, 40, 18], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════
# Sheet 5: 코드 규모 분석
# ═══════════════════════════════════════
ws5 = wb.create_sheet('코드 규모 분석')
ws5.sheet_properties.tabColor = 'FFC000'
add_title(ws5, '코드 규모 분석 (LOC 기반)', merge_to=6)

headers5 = ['구분', '분류', '파일 수', 'LOC', '평균 LOC/파일', '설명']
r = 3
for c, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, 6)

loc_data = [
    ('프론트엔드', '컴포넌트 (React)', 163, 28319, None, 'UI/CRUD/차트/캔버스/네트워크'),
    ('', '페이지 (page.tsx)', 34, 7466, None, '라우트별 페이지'),
    ('', '차트 옵션·설정', 15, 3255, None, 'ECharts 옵션 빌더'),
    ('', 'API 클라이언트', 27, 2548, None, 'fetch 래퍼 + 타입 안전'),
    ('', '타입 정의', 19, 1891, None, 'TypeScript 인터페이스'),
    ('', '상태 관리 (Zustand)', 17, 1712, None, '17개 스토어'),
    ('', '커스텀 훅', 11, 1537, None, '재사용 로직'),
    ('', '유틸리티·기타', 50, 4269, None, '포매터, 상수, 설정'),
    ('프론트엔드 소계', '', 336, 50997, None, ''),
    ('백엔드', 'ai_server.py (메인)', 1, 14846, None, '106+ 엔드포인트'),
    ('', '이상감지 모듈', 3, 2926, None, 'anomaly_detector + iforest + profiler'),
    ('', '인텐트 분류', 4, 1951, None, 'classifier + embeddings + index + extractor'),
    ('', '물 수지 검증', 1, 443, None, 'flow_balance.py'),
    ('', 'SNMP 폴링', 1, 414, None, 'snmp_poller.py'),
    ('', 'DB 동기화', 2, 524, None, 'db_sync + sync_timeseries'),
    ('', '기타 모듈', 12, 2071, None, '세션/Ollama/퍼지/설정/테스트'),
    ('백엔드 소계', '', 24, 23175, None, ''),
    ('전체 합계', '', 360, 74172, None, ''),
]

for i, d in enumerate(loc_data):
    r = 4 + i
    for c, v in enumerate(d, 1):
        ws5.cell(row=r, column=c, value=v)
    if d[4] is None and d[2] and d[3]:
        ws5.cell(row=r, column=5, value=f'=IF(C{r}>0,ROUND(D{r}/C{r},0),"")')
    aligns = [LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, RIGHT, LEFT_WRAP]
    is_total = '소계' in str(d[0]) or '합계' in str(d[0])
    style_row(ws5, r, 6, fill=TOTAL_FILL if is_total else (CATEGORY_FILL if d[0] and not is_total else None),
              font=BOLD_FONT if (d[0] or is_total) else NORMAL_FONT, aligns=aligns)
    ws5.cell(row=r, column=3).number_format = NUM_FMT
    ws5.cell(row=r, column=4).number_format = NUM_FMT
    ws5.cell(row=r, column=5).number_format = NUM_FMT

for c, w in enumerate([16, 25, 10, 10, 14, 35], 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════
# Sheet 6: 투입 인력 산정
# ═══════════════════════════════════════
ws6 = wb.create_sheet('투입 인력 산정')
ws6.sheet_properties.tabColor = '7030A0'
add_title(ws6, '투입 인력 산정 (M/M 기반)', merge_to=7)

headers6 = ['No', '업무 영역', '세부 작업', '등급', '투입 M/M', '단가(원/월)', '금액(원)']
r = 3
for c, h in enumerate(headers6, 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, 7)

# 한국 SW기술자 노임단가 2026 기준 (예시)
mm_data = [
    ('요구사항 분석', '시스템 요구사항 분석 및 설계', '특급기술자', 1.5, 13357530, None),
    ('', '화면 설계 (34개 페이지)', '고급기술자', 1.0, 10624670, None),
    ('', 'DB 설계 (50테이블+7뷰)', '특급기술자', 1.0, 13357530, None),
    ('', 'API 설계 (104개 엔드포인트)', '특급기술자', 0.5, 13357530, None),
    ('프론트엔드 개발', '페이지 개발 (34개)', '고급기술자', 3.0, 10624670, None),
    ('', '컴포넌트 개발 (163개)', '고급기술자', 4.0, 10624670, None),
    ('', '차트/시각화 (14종)', '특급기술자', 2.5, 13357530, None),
    ('', '캔버스 에디터 (React Flow)', '특급기술자', 1.5, 13357530, None),
    ('', '용수 흐름 모니터링 (SVG)', '특급기술자', 2.0, 13357530, None),
    ('', '상태관리+훅+API클라이언트', '중급기술자', 2.0, 8389890, None),
    ('백엔드 개발', 'FastAPI 서버 (106+ API)', '특급기술자', 4.0, 13357530, None),
    ('', 'AI 인텐트 분류 (68 인텐트)', '특급기술자', 2.0, 13357530, None),
    ('', '이상감지 엔진 (4계층)', '특급기술자', 3.0, 13357530, None),
    ('', '인과관계 엔진 (Rule)', '특급기술자', 2.0, 13357530, None),
    ('', '물 수지 검증 + 교차검증', '고급기술자', 1.5, 10624670, None),
    ('', 'SNMP + 네트워크 폴링', '고급기술자', 1.0, 10624670, None),
    ('DB 구축', '테이블 생성 + 인덱스', '중급기술자', 1.5, 8389890, None),
    ('', '데이터 마이그레이션', '중급기술자', 1.0, 8389890, None),
    ('', 'TimescaleDB 최적화', '고급기술자', 1.0, 10624670, None),
    ('', '함수/뷰/집계 개발 (20개)', '고급기술자', 1.0, 10624670, None),
    ('테스트·최적화', '단위 테스트 + E2E 테스트', '중급기술자', 2.0, 8389890, None),
    ('', '성능 최적화 (쿼리+캐시)', '특급기술자', 1.5, 13357530, None),
    ('', '통합 테스트 + 검증', '고급기술자', 1.0, 10624670, None),
    ('인프라·배포', 'Docker + HTTPS + Tunnel', '중급기술자', 0.5, 8389890, None),
    ('PM·관리', '프로젝트 관리', '특급기술자', 2.0, 13357530, None),
]

for i, d in enumerate(mm_data):
    r = 4 + i
    ws6.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(d, 2):
        ws6.cell(row=r, column=c, value=v)
    ws6.cell(row=r, column=7, value=f'=E{r}*F{r}')
    ws6.cell(row=r, column=7).number_format = MONEY_FMT
    ws6.cell(row=r, column=6).number_format = MONEY_FMT
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, CENTER, RIGHT, RIGHT, RIGHT]
    style_row(ws6, r, 7, fill=CATEGORY_FILL if d[0] else None, font=BOLD_FONT if d[0] else NORMAL_FONT, aligns=aligns)

# 소계 행
r_mm_total = 4 + len(mm_data)
ws6.cell(row=r_mm_total, column=1, value='합계')
ws6.merge_cells(start_row=r_mm_total, start_column=1, end_row=r_mm_total, end_column=4)
ws6.cell(row=r_mm_total, column=5, value=f'=SUM(E4:E{r_mm_total-1})')
ws6.cell(row=r_mm_total, column=5).number_format = '0.0'
ws6.cell(row=r_mm_total, column=7, value=f'=SUM(G4:G{r_mm_total-1})')
ws6.cell(row=r_mm_total, column=7).number_format = MONEY_FMT
style_row(ws6, r_mm_total, 7, fill=TOTAL_FILL, font=Font(name='맑은 고딕', bold=True, size=11))
ws6.cell(row=r_mm_total, column=1).alignment = LEFT_WRAP
ws6.cell(row=r_mm_total, column=5).alignment = RIGHT
ws6.cell(row=r_mm_total, column=7).alignment = RIGHT

# 직접인건비 / 제경비 / 기술료
r_e = r_mm_total
r_e += 1
ws6.cell(row=r_e, column=2, value='직접인건비 (위 합계)')
ws6.cell(row=r_e, column=7, value=f'=G{r_mm_total}')
ws6.cell(row=r_e, column=7).number_format = MONEY_FMT
style_row(ws6, r_e, 7, font=BOLD_FONT)
ws6.cell(row=r_e, column=2).alignment = LEFT_WRAP
ws6.cell(row=r_e, column=7).alignment = RIGHT

r_e += 1
ws6.cell(row=r_e, column=2, value='제경비 (직접인건비 × 110%)')
ws6.cell(row=r_e, column=7, value=f'=G{r_e-1}*1.1')
ws6.cell(row=r_e, column=7).number_format = MONEY_FMT
style_row(ws6, r_e, 7, font=NORMAL_FONT)
ws6.cell(row=r_e, column=2).alignment = LEFT_WRAP
ws6.cell(row=r_e, column=7).alignment = RIGHT

r_e += 1
ws6.cell(row=r_e, column=2, value='기술료 ((직접인건비+제경비) × 20%)')
ws6.cell(row=r_e, column=7, value=f'=(G{r_e-2}+G{r_e-1})*0.2')
ws6.cell(row=r_e, column=7).number_format = MONEY_FMT
style_row(ws6, r_e, 7, font=NORMAL_FONT)
ws6.cell(row=r_e, column=2).alignment = LEFT_WRAP
ws6.cell(row=r_e, column=7).alignment = RIGHT

r_e += 1
ws6.cell(row=r_e, column=2, value='소프트웨어 개발비 소계')
ws6.cell(row=r_e, column=7, value=f'=G{r_e-3}+G{r_e-2}+G{r_e-1}')
ws6.cell(row=r_e, column=7).number_format = MONEY_FMT
style_row(ws6, r_e, 7, fill=SUBTOTAL_FILL, font=BOLD_FONT)
ws6.cell(row=r_e, column=2).alignment = LEFT_WRAP
ws6.cell(row=r_e, column=7).alignment = RIGHT

r_e += 1
ws6.cell(row=r_e, column=2, value='부가가치세 (10%)')
ws6.cell(row=r_e, column=7, value=f'=G{r_e-1}*0.1')
ws6.cell(row=r_e, column=7).number_format = MONEY_FMT
style_row(ws6, r_e, 7, font=NORMAL_FONT)
ws6.cell(row=r_e, column=2).alignment = LEFT_WRAP
ws6.cell(row=r_e, column=7).alignment = RIGHT

r_e += 1
ws6.cell(row=r_e, column=2, value='총 사업비 (VAT 포함)')
ws6.cell(row=r_e, column=7, value=f'=G{r_e-2}+G{r_e-1}')
ws6.cell(row=r_e, column=7).number_format = MONEY_FMT
style_row(ws6, r_e, 7, fill=PatternFill('solid', fgColor='FFD966'), font=Font(name='맑은 고딕', bold=True, size=12))
ws6.cell(row=r_e, column=2).alignment = LEFT_WRAP
ws6.cell(row=r_e, column=7).alignment = RIGHT

for c, w in enumerate([5, 18, 32, 12, 10, 16, 18], 1):
    ws6.column_dimensions[get_column_letter(c)].width = w

# ── 인쇄 설정 ──
for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

OUTPUT = 'D:/web/SLM_대가산정표.xlsx'
wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
