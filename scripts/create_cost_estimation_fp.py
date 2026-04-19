"""
SW사업 대가산정 가이드 (2024년 개정) 기능점수(FP) 방식
- 과학기술정보통신부 / 한국소프트웨어산업협회(KOSA)
- 기능점수 산정 → 보정계수 적용 → 대가 산출
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 공통 스타일 ──
HEADER_FONT = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
SUB_HEADER_FILL = PatternFill('solid', fgColor='4472C4')
CATEGORY_FILL = PatternFill('solid', fgColor='D6E4F0')
LIGHT_BLUE = PatternFill('solid', fgColor='DAEEF3')
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14)
SUBTITLE_FONT = Font(name='맑은 고딕', bold=True, size=12)
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=11, color='2F5496')
NORMAL_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', bold=True, size=10)
SMALL_FONT = Font(name='맑은 고딕', size=9, color='666666')
TOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')
SUBTOTAL_FILL = PatternFill('solid', fgColor='E2EFDA')
WARN_FILL = PatternFill('solid', fgColor='FCE4EC')
BLUE_TEXT = Font(name='맑은 고딕', size=10, color='0000FF')
thin = Side(style='thin', color='B4C6E7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
NUM_FMT = '#,##0'
MONEY_FMT = '#,##0'
PCT_FMT = '0.00'

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_sub_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
        cell.fill = SUB_HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_row(ws, row, cols, fill=None, font=None, aligns=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or NORMAL_FONT
        cell.border = BORDER
        if fill: cell.fill = fill
        if aligns and c <= len(aligns):
            cell.alignment = aligns[c-1]
        else:
            cell.alignment = CENTER

def add_title(ws, title, row=1, col=1, merge_to=8):
    ws.cell(row=row, column=col, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)

def set_widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════════════════════════════
# Sheet 1: 대가산정 총괄표 (공공SW사업 기능점수 방식)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '대가산정 총괄표'
ws1.sheet_properties.tabColor = '2F5496'

add_title(ws1, 'SLM 스마트 관망관리 시스템 — SW개발비 대가산정 총괄표')
ws1.cell(row=2, column=1, value='근거: SW사업 대가산정 가이드(2024년 개정) — 기능점수(FP) 방식').font = SMALL_FONT
ws1.merge_cells('A2:H2')
ws1.cell(row=3, column=1, value='작성일: 2026-03-14 | 발주기관: 한국수자원공사(K-water)').font = SMALL_FONT
ws1.merge_cells('A3:H3')

# ── 산출 구조 ──
r = 5
headers = ['구분', '항목', '산출 근거', '단위', '수량', '단가(원)', '금액(원)', '비고']
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 8)

rows_data = [
    # Section 1: 기능점수 기반 개발비
    ('Ⅰ. SW개발비', '1. 보정 전 개발원가', '기능점수(FP) × FP단가', 'FP', '', '', '', '「기능점수 산정」시트 참조'),
    ('', '  (1) 기능점수(FP)', '「기능점수 산정」시트 합계', 'FP', None, '', '', 'ILF+EIF+EI+EO+EQ'),  # will be formula
    ('', '  (2) FP 단가', '2025년 고시 단가', '원/FP', 1, 553114, None, 'KOSA 공표'),
    ('', '  (3) 보정 전 개발원가', '(1) × (2)', '원', '', '', None, ''),  # formula
    ('', '2. 보정계수', '「보정계수」시트 참조', '', '', '', '', ''),
    ('', '  (1) 규모보정계수', 'FP 규모에 따른 보정', '', '', '', '', '「보정계수」시트'),
    ('', '  (2) 연계복잡성수준', '외부 연계 시스템 수', '', '', '', '', ''),
    ('', '  (3) 성능요구수준', '응답시간·처리량 요구', '', '', '', '', ''),
    ('', '  (4) 운영환경호환성', '호환성 요구사항', '', '', '', '', ''),
    ('', '  (5) 보안수준', '보안 요구사항', '', '', '', '', ''),
    ('', '  (6) 종합 보정계수', '(1)×(2)×(3)×(4)×(5)', '', '', '', None, '', ''),  # formula
    ('', '3. 보정 후 개발원가', '보정 전 원가 × 종합 보정계수', '원', '', '', None, ''),  # formula
    ('', '4. 직접경비', '여비, 재료비, 인쇄비, 교육훈련비 등', '원', 1, '', None, '직접경비 산정 기준'),
    ('', '5. SW개발비 합계', '보정 후 개발원가 + 직접경비', '원', '', '', None, ''),
    ('Ⅱ. 부가가치세', 'VAT (10%)', 'SW개발비 × 10%', '원', '', '', None, ''),
    ('Ⅲ. 총 사업비', '합계 (VAT 포함)', 'SW개발비 + VAT', '원', '', '', None, ''),
]

aligns8 = [LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, CENTER, RIGHT, RIGHT, RIGHT, LEFT_WRAP]

for i, d in enumerate(rows_data):
    r = 6 + i
    for c, v in enumerate(d, 1):
        if v is not None:
            ws1.cell(row=r, column=c, value=v)
    is_section = bool(d[0]) and d[0].startswith('Ⅰ') or d[0].startswith('Ⅱ') or d[0].startswith('Ⅲ')
    fill = CATEGORY_FILL if is_section else None
    font = BOLD_FONT if (is_section or '합계' in str(d[1])) else NORMAL_FONT
    style_row(ws1, r, 8, fill=fill, font=font, aligns=aligns8)
    for cc in [5, 6, 7]:
        ws1.cell(row=r, column=cc).number_format = MONEY_FMT

# ── 수식 연결 (기능점수 산정 시트 참조) ──
# Row references
R_FP = 7       # (1) 기능점수
R_UNIT = 8     # (2) FP단가
R_RAW = 9      # (3) 보정 전 개발원가
R_CORR = 16    # (6) 종합 보정계수
R_ADJ = 17     # 3. 보정 후 개발원가
R_DIRECT = 18  # 4. 직접경비
R_TOTAL = 19   # 5. SW개발비 합계
R_VAT = 20     # VAT
R_GRAND = 21   # 총 사업비

ws1.cell(row=R_FP, column=5, value=f"='기능점수 산정'!H60")  # FP합계 셀 참조
ws1.cell(row=R_FP, column=5).font = BLUE_TEXT
ws1.cell(row=R_RAW, column=7, value=f'=E{R_FP}*F{R_UNIT}')
ws1.cell(row=R_CORR, column=7, value=f"='보정계수'!D12")
ws1.cell(row=R_CORR, column=7).font = BLUE_TEXT
ws1.cell(row=R_ADJ, column=7, value=f'=G{R_RAW}*G{R_CORR}')
ws1.cell(row=R_ADJ, column=7).number_format = MONEY_FMT
ws1.cell(row=R_DIRECT, column=6, value=5000000)
ws1.cell(row=R_DIRECT, column=7, value=f'=F{R_DIRECT}')
ws1.cell(row=R_TOTAL, column=7, value=f'=G{R_ADJ}+G{R_DIRECT}')
style_row(ws1, R_TOTAL, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT, aligns=aligns8)
ws1.cell(row=R_VAT, column=7, value=f'=G{R_TOTAL}*0.1')
ws1.cell(row=R_GRAND, column=7, value=f'=G{R_TOTAL}+G{R_VAT}')
style_row(ws1, R_GRAND, 8, fill=PatternFill('solid', fgColor='FFD966'),
          font=Font(name='맑은 고딕', bold=True, size=12), aligns=aligns8)
for rr in [R_RAW, R_ADJ, R_DIRECT, R_TOTAL, R_VAT, R_GRAND]:
    ws1.cell(row=rr, column=7).number_format = MONEY_FMT

# 보정계수 참조 행
for rr, ref in [(11, "='보정계수'!D3"), (12, "='보정계수'!D5"), (13, "='보정계수'!D7"),
                (14, "='보정계수'!D9"), (15, "='보정계수'!D11")]:
    ws1.cell(row=rr, column=7, value=ref)
    ws1.cell(row=rr, column=7).font = BLUE_TEXT
    ws1.cell(row=rr, column=7).number_format = PCT_FMT

set_widths(ws1, [14, 22, 30, 8, 12, 14, 18, 25])

# ═══════════════════════════════════════════════════════════════
# Sheet 2: 기능점수 산정
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('기능점수 산정')
ws2.sheet_properties.tabColor = '4472C4'
add_title(ws2, '기능점수(Function Point) 산정 내역', merge_to=8)
ws2.cell(row=2, column=1, value='산정 기준: IFPUG CPM 4.3.1 / 한국소프트웨어산업협회 간이법').font = SMALL_FONT
ws2.merge_cells('A2:H2')

# ── ILF (내부논리파일) ──
r = 4
ws2.cell(row=r, column=1, value='1. 데이터 기능 — ILF (Internal Logical File)').font = SECTION_FONT
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r = 5
headers_fp = ['No', '기능명', '관련 테이블', 'DET', 'RET', '복잡도', 'FP', '비고']
for c, h in enumerate(headers_fp, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 8)

# ILF: Low=7, Avg=10, High=15
# DET=데이터요소유형(컬럼수), RET=레코드요소유형(서브그룹수)
ilf_data = [
    ('사용자/인증 관리', 'tb_user, tb_auth, tb_user_session, tb_access_log', 28, 4, '보통', 10, ''),
    ('메뉴/권한 관리', 'tb_menu, tb_auth_menu, tb_menu_api', 18, 3, '단순', 7, '자기참조 FK'),
    ('공통코드', 'tb_grp_code, tb_comm_code', 8, 2, '단순', 7, ''),
    ('AI 채팅 대화', 'tb_ai_chat_ask_group, ask, bot, ask_image, bot_image', 35, 5, '복잡', 15, 'visual_data JSONB'),
    ('AI FAQ', 'tb_ai_chat_faq', 6, 1, '단순', 7, ''),
    ('배수지 정보', 'tb_service_reservoir_info, _status', 45, 3, '보통', 10, 'general_overview JSONB'),
    ('가압장 정보', 'tb_service_booster_station_info, _status', 38, 3, '보통', 10, '설비메타 26항목'),
    ('감압시설 정보', 'tb_pressure_reducing_facility_info, _status', 22, 2, '단순', 7, ''),
    ('블록 정보', 'tb_block_info, _status', 20, 2, '단순', 7, 'block_level 구분'),
    ('설비 마스터', 'tb_equipment_info, _status, _alarm_report', 42, 3, '복잡', 15, '290건'),
    ('설비↔태그 매핑', 'tb_equipment_tag_map', 5, 1, '단순', 7, '3,375건 자동매핑'),
    ('네트워크 연결', 'tb_network_link', 8, 1, '단순', 7, '방향성 엣지'),
    ('용수 흐름 계통도', 'tb_facility_flow_map', 8, 1, '단순', 7, '95엣지'),
    ('태그 마스터', 'tb_tag_info', 15, 1, '단순', 7, '2,698건'),
    ('태그 시계열', 'tb_tag_raw_data', 5, 1, '단순', 7, 'TimescaleDB 하이퍼테이블'),
    ('태그 분류 체계', 'tb_tag_data_group, tb_tag_group_map', 12, 2, '단순', 7, '21그룹, 93% 자동'),
    ('트렌드 카탈로그', 'tb_trend_catalog', 8, 1, '단순', 7, ''),
    ('모니터링 카탈로그', 'tb_monitoring_catalog', 10, 1, '단순', 7, ''),
    ('알람 이력', 'tb_alarm_log', 15, 1, '단순', 7, 'HH/LL/H/L'),
    ('현장 프로파일', 'tb_site_anomaly_profile', 12, 1, '단순', 7, 'A/B/C/D 그룹'),
    ('캔버스 노드 위치', 'tb_canvas_node_position', 6, 1, '단순', 7, ''),
    ('인과 규칙 오버라이드', 'tb_causal_chain_override', 8, 1, '보통', 10, 'JSONB chain'),
    ('SNMP 포트 상태', 'tb_snmp_port_status', 12, 1, '단순', 7, '13대×24포트'),
    ('프롬프트 템플릿', 'tb_prompt_template, _column', 12, 2, '단순', 7, ''),
    ('파일 관리', 'tb_file_storage, _history, tb_facility_file', 20, 3, '보통', 10, ''),
    ('필드 잠금', 'tb_field_lock', 5, 1, '단순', 7, ''),
    ('사이트 설정', 'tb_admin_site_settings', 4, 1, '단순', 7, ''),
]

for i, d in enumerate(ilf_data):
    r = 6 + i
    ws2.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(d, 2):
        ws2.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, CENTER, RIGHT, LEFT_WRAP]
    style_row(ws2, r, 8, aligns=aligns)

r_ilf_sub = 6 + len(ilf_data)
ws2.cell(row=r_ilf_sub, column=1, value='ILF 소계')
ws2.merge_cells(start_row=r_ilf_sub, start_column=1, end_row=r_ilf_sub, end_column=6)
ws2.cell(row=r_ilf_sub, column=7, value=f'=SUM(G6:G{r_ilf_sub-1})')
ws2.cell(row=r_ilf_sub, column=8, value=f'{len(ilf_data)}개 ILF')
style_row(ws2, r_ilf_sub, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT)

# ── EIF (외부인터페이스파일) ──
r = r_ilf_sub + 2
r_eif_title = r
ws2.cell(row=r, column=1, value='2. 데이터 기능 — EIF (External Interface File)').font = SECTION_FONT
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
for c, h in enumerate(headers_fp, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 8)

# EIF: Low=5, Avg=7, High=10
eif_data = [
    ('원격 운영 DB', '112.166.183.65:25479 PostgreSQL', 50, 5, '복잡', 10, '시계열 동기화'),
    ('Ollama SLM API', 'Phi-4-mini, snowflake-arctic-embed2', 8, 2, '단순', 5, '인텐트 분류 폴백'),
    ('SNMP 네트워크 장비', '13대 L2/L3 스위치 (24포트)', 12, 2, '보통', 7, 'Mock/Real 듀얼'),
    ('Node-RED 경보분석', 'Docker :1880, 경보 진단 메시지', 6, 1, '단순', 5, 'iframe 연동'),
]

r_eif_start = r + 1
for i, d in enumerate(eif_data):
    r = r_eif_start + i
    ws2.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(d, 2):
        ws2.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, CENTER, RIGHT, LEFT_WRAP]
    style_row(ws2, r, 8, aligns=aligns)

r_eif_sub = r_eif_start + len(eif_data)
ws2.cell(row=r_eif_sub, column=1, value='EIF 소계')
ws2.merge_cells(start_row=r_eif_sub, start_column=1, end_row=r_eif_sub, end_column=6)
ws2.cell(row=r_eif_sub, column=7, value=f'=SUM(G{r_eif_start}:G{r_eif_sub-1})')
ws2.cell(row=r_eif_sub, column=8, value=f'{len(eif_data)}개 EIF')
style_row(ws2, r_eif_sub, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT)

# ── EI (외부입력) ──
r = r_eif_sub + 2
ws2.cell(row=r, column=1, value='3. 트랜잭션 기능 — EI (External Input)').font = SECTION_FONT
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
for c, h in enumerate(headers_fp, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 8)

# EI: Low=3, Avg=4, High=6
ei_data = [
    ('사용자 로그인', 'JWT 인증, bcrypt 검증', 6, 1, '단순', 3, 'NextAuth'),
    ('AI 채팅 질의 (동기)', '인텐트 분류→SQL→응답 생성', 15, 3, '복잡', 6, '68인텐트'),
    ('AI 채팅 질의 (SSE)', 'SSE 스트리밍 (token→visual→done)', 15, 3, '복잡', 6, '실시간'),
    ('AI 인텐트 보정', 'force_intent 파라미터', 5, 1, '단순', 3, '오분류 사후보정'),
    ('모니터링 카탈로그 생성', 'POST /monitoring/catalogs', 10, 1, '보통', 4, ''),
    ('모니터링 카탈로그 수정', 'PUT /monitoring/catalogs/{id}', 10, 1, '보통', 4, ''),
    ('모니터링 카탈로그 삭제', 'DELETE /monitoring/catalogs/{id}', 2, 1, '단순', 3, ''),
    ('설비 등록', 'POST /equipments (자동 ID 생성)', 15, 1, '보통', 4, ''),
    ('설비 수정', 'PUT /equipments/{id}', 15, 1, '보통', 4, ''),
    ('설비 삭제', 'DELETE /equipments/{id} (CASCADE 체크)', 3, 1, '단순', 3, 'dry_run'),
    ('설비↔태그 자동 매핑', 'GET /equipments/auto-map', 8, 2, '보통', 4, '3,375건'),
    ('설비↔태그 수동 연결/해제', 'POST/DELETE equipment-tag-link', 4, 1, '단순', 3, '캔버스'),
    ('배수지 등록', 'POST /reservoirs (info+status)', 27, 3, '복잡', 6, '3탭 폼'),
    ('배수지 수정', 'PUT /reservoirs/{sn}', 27, 3, '복잡', 6, ''),
    ('배수지 삭제', 'DELETE /reservoirs/{sn}', 2, 1, '단순', 3, 'dry_run'),
    ('가압장 등록/수정', 'POST/PUT /boosters', 18, 2, '보통', 4, '2건 합산'),
    ('가압장 삭제', 'DELETE /boosters/{sn}', 2, 1, '단순', 3, ''),
    ('감압시설 등록/수정', 'POST/PUT /pressure-reducing', 9, 2, '단순', 3, '2건 합산'),
    ('감압시설 삭제', 'DELETE /pressure-reducing/{sn}', 2, 1, '단순', 3, ''),
    ('블록 등록/수정', 'POST/PUT /blocks', 11, 2, '보통', 4, '2건 합산'),
    ('블록 삭제', 'DELETE /blocks/{sn}', 2, 1, '단순', 3, ''),
    ('용수 흐름 연결 추가/삭제', 'POST/DELETE /flow-map', 6, 1, '단순', 3, ''),
    ('네트워크 장비 CRUD', 'POST/PUT/DELETE /network/infos', 12, 1, '보통', 4, '3건 합산'),
    ('네트워크 연결 CRUD', 'POST/PUT/DELETE /network/links', 8, 1, '보통', 4, '3건 합산'),
    ('경보 확인 처리', 'PUT /crisis/alarm-reports/confirm', 3, 1, '단순', 3, ''),
    ('캔버스 레이아웃 저장', 'PUT /canvas/layout (UPSERT+diff)', 20, 2, '복잡', 6, '엣지 diff'),
    ('인과 규칙 저장/삭제', 'PUT/DELETE /causal/chain', 12, 2, '보통', 4, 'JSONB'),
    ('시설 이미지 업로드', 'POST /admin/facility-files/upload', 8, 1, '보통', 4, 'multipart'),
    ('시설 파일 삭제', 'DELETE /admin/facility-files/{id}', 2, 1, '단순', 3, ''),
    ('사이트 설정 수정', 'PUT /admin/site-settings', 5, 1, '단순', 3, ''),
    ('CSV 일괄 등록 (8종)', 'POST /*/import/csv (태그/설비/시설 8종)', 20, 2, '복잡', 6, '8개 메뉴'),
    ('Ollama 모델 변경', 'POST /models/select', 3, 1, '단순', 3, ''),
]

r_ei_start = r + 1
for i, d in enumerate(ei_data):
    r = r_ei_start + i
    ws2.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(d, 2):
        ws2.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, CENTER, RIGHT, LEFT_WRAP]
    style_row(ws2, r, 8, aligns=aligns)

r_ei_sub = r_ei_start + len(ei_data)
ws2.cell(row=r_ei_sub, column=1, value='EI 소계')
ws2.merge_cells(start_row=r_ei_sub, start_column=1, end_row=r_ei_sub, end_column=6)
ws2.cell(row=r_ei_sub, column=7, value=f'=SUM(G{r_ei_start}:G{r_ei_sub-1})')
ws2.cell(row=r_ei_sub, column=8, value=f'{len(ei_data)}개 EI')
style_row(ws2, r_ei_sub, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT)

# ── EO (외부출력) ──
r = r_ei_sub + 2
ws2.cell(row=r, column=1, value='4. 트랜잭션 기능 — EO (External Output)').font = SECTION_FONT
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
for c, h in enumerate(headers_fp, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 8)

# EO: Low=4, Avg=5, High=7
eo_data = [
    ('종합 현황판', 'GET /dashboard/overview (6종 KPI 캐시 집계)', 35, 4, '복잡', 7, '팝업 분석 연결'),
    ('용수 흐름 실시간', 'GET /flow-map/realtime (유량+교차+물수지+장애+공급시간)', 50, 6, '복잡', 7, '핵심 기능'),
    ('이상감지 전체 스캔', 'ANOMALY_SCAN_ALL (4계층+교차+물수지+품질+장애)', 45, 5, '복잡', 7, '캐시 사전계산'),
    ('이상감지 시설 상세', 'ANOMALY_FACILITY_DETAIL (인과+교차+전파+내부)', 40, 5, '복잡', 7, 'Rule 엔진'),
    ('시계열 트렌드', 'POST /trend/data (청크 직접쿼리, 15태그)', 20, 2, '보통', 5, 'TimescaleDB'),
    ('야간최소유량 표', '_execute_night_min_flow_query (60분 이동평균)', 15, 2, '보통', 5, '청크 최적화'),
    ('야간최소유량 표준편차', '_execute_night_min_flow_stddev_query (400일 통계)', 20, 2, '복잡', 7, '다중 시설 뷰'),
    ('결측분석', '_execute_tag_daily_summary_query (SQL 분단위 집계)', 15, 2, '보통', 5, '청크 최적화'),
    ('카탈로그 트렌드 표', '_execute_catalog_trend_query (2단계 청크)', 15, 2, '보통', 5, ''),
    ('교차시설 검증', 'ANOMALY_CROSS_FACILITY (상류→하류 일관성)', 20, 3, '복잡', 7, 'BFS 탐색'),
    ('물 수지 검증', 'ANOMALY_FLOW_BALANCE (유량적분, 5등급)', 25, 3, '복잡', 7, '사다리꼴 적분'),
    ('경보 이상지점 조회', 'ALARM_ABNORMAL_LOCATIONS (7일 폴백)', 12, 2, '보통', 5, ''),
    ('경보분석 상세', 'GET /crisis/alarm-analysis/detail', 15, 2, '보통', 5, '진단 메시지'),
    ('경보 대시보드', 'GET /crisis/alarm-dashboard', 12, 2, '보통', 5, '도넛+카테고리'),
    ('네트워크 토폴로지', 'GET /network/topology (노드+엣지)', 15, 2, '보통', 5, 'Force+계층형'),
    ('SNMP 포트 요약', 'GET /network/snmp/summary', 10, 1, '단순', 4, '13대'),
    ('인과 규칙 현황', 'GET /causal/rules (템플릿+커버리지+태그매핑)', 20, 3, '복잡', 7, '5 시설유형'),
    ('인과 시간지연 추정', 'POST /causal/estimate-lag (교차상관)', 8, 1, '보통', 5, 'numpy'),
    ('현장 프로파일', 'GET /anomaly/profiles (A/B/C/D)', 12, 1, '보통', 5, ''),
    ('자동완성', 'GET /autocomplete/candidates', 8, 1, '단순', 4, ''),
    ('알람 알림', 'GET /monitoring/alarm-notifications', 6, 1, '단순', 4, '헤더 벨'),
    ('모니터링 대시보드', 'GET /monitoring/dashboard', 15, 2, '보통', 5, ''),
    ('CSV 내보내기 (8종)', 'fetchAllPages + downloadCsv (빈 템플릿 포함)', 15, 1, '보통', 5, '8개 메뉴'),
]

r_eo_start = r + 1
for i, d in enumerate(eo_data):
    r = r_eo_start + i
    ws2.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(d, 2):
        ws2.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, CENTER, RIGHT, LEFT_WRAP]
    style_row(ws2, r, 8, aligns=aligns)

r_eo_sub = r_eo_start + len(eo_data)
ws2.cell(row=r_eo_sub, column=1, value='EO 소계')
ws2.merge_cells(start_row=r_eo_sub, start_column=1, end_row=r_eo_sub, end_column=6)
ws2.cell(row=r_eo_sub, column=7, value=f'=SUM(G{r_eo_start}:G{r_eo_sub-1})')
ws2.cell(row=r_eo_sub, column=8, value=f'{len(eo_data)}개 EO')
style_row(ws2, r_eo_sub, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT)

# ── EQ (외부조회) ──
r = r_eo_sub + 2
ws2.cell(row=r, column=1, value='5. 트랜잭션 기능 — EQ (External Inquiry)').font = SECTION_FONT
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
for c, h in enumerate(headers_fp, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 8)

# EQ: Low=3, Avg=4, High=6
eq_data = [
    ('태그 목록 조회', 'GET /tags (5종 필터+페이징)', 15, 1, '보통', 4, '2,698건'),
    ('태그 필터 옵션', 'GET /tags/filters', 5, 1, '단순', 3, ''),
    ('태그 그룹 통계', 'GET /tags/groups (21그룹)', 10, 1, '보통', 4, ''),
    ('설비 목록 조회', 'GET /equipments (3종 필터+페이징)', 12, 1, '보통', 4, ''),
    ('설비 필터/다음ID', 'GET /equipments/filters, next-id', 5, 1, '단순', 3, '2건'),
    ('배수지 목록/상세', 'GET /reservoirs, /{sn}', 25, 2, '보통', 4, '2건'),
    ('가압장 목록/상세', 'GET /boosters, /{sn}', 20, 2, '보통', 4, '2건'),
    ('감압시설 목록/상세', 'GET /pressure-reducing, /{sn}', 12, 2, '단순', 3, '2건'),
    ('블록 목록/상세', 'GET /blocks, /{sn}', 12, 2, '단순', 3, '2건'),
    ('용수 흐름 계통도', 'GET /flow-map (토폴로지+메타)', 10, 1, '보통', 4, '95엣지'),
    ('용수 흐름 루트/하류', 'GET /flow-map/roots, /downstream', 8, 1, '단순', 3, '2건'),
    ('용수 흐름 노드 알람', 'GET /flow-map/node-alarms', 8, 1, '단순', 3, ''),
    ('네트워크 장비 조회', 'GET /network/devices', 15, 1, '보통', 4, ''),
    ('네트워크 정보/필터', 'GET /network/infos, /filters', 12, 1, '보통', 4, '2건'),
    ('네트워크 연결 조회', 'GET /network/links, /protocols', 10, 1, '단순', 3, '2건'),
    ('네트워크 장비 검색', 'GET /network/links/equipment-search', 6, 1, '단순', 3, ''),
    ('네트워크 상태 요약', 'GET /network/status/summary', 10, 1, '보통', 4, ''),
    ('SNMP 포트/시스템', 'GET /snmp/{id}/ports, /system', 12, 1, '보통', 4, '2건'),
    ('모니터링 카탈로그', 'GET /monitoring/catalogs, /sites, /reference', 10, 1, '보통', 4, '3건'),
    ('경보 리포트', 'GET /crisis/alarm-reports', 12, 1, '보통', 4, ''),
    ('캔버스 레이아웃', 'GET /canvas/layout', 15, 2, '보통', 4, '78노드'),
    ('캔버스 노드 상세', 'GET /canvas/node-detail/{sn}/{ft}', 20, 3, '복잡', 6, '설비+카탈로그'),
    ('캔버스 설비 태그', 'GET /canvas/equipment-tags/{sn}/{ft}', 8, 1, '단순', 3, ''),
    ('인과 체인 조회', 'GET /causal/chain/{sn}/{ft}', 12, 2, '보통', 4, ''),
    ('인과 검증 디버그', 'GET /causal/verify', 10, 1, '보통', 4, ''),
    ('관리자 파일/현황', 'GET /admin/facility-files, /facilities-summary', 12, 2, '보통', 4, '2건'),
    ('사이트 설정 조회', 'GET /admin/site-settings', 4, 1, '단순', 3, ''),
    ('서버 상태', 'GET /health, /models', 6, 1, '단순', 3, '2건'),
    ('CSV 다운로드', 'GET /csv/{filename}', 3, 1, '단순', 3, ''),
]

r_eq_start = r + 1
for i, d in enumerate(eq_data):
    r = r_eq_start + i
    ws2.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(d, 2):
        ws2.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, CENTER, RIGHT, LEFT_WRAP]
    style_row(ws2, r, 8, aligns=aligns)

r_eq_sub = r_eq_start + len(eq_data)
ws2.cell(row=r_eq_sub, column=1, value='EQ 소계')
ws2.merge_cells(start_row=r_eq_sub, start_column=1, end_row=r_eq_sub, end_column=6)
ws2.cell(row=r_eq_sub, column=7, value=f'=SUM(G{r_eq_start}:G{r_eq_sub-1})')
ws2.cell(row=r_eq_sub, column=8, value=f'{len(eq_data)}개 EQ')
style_row(ws2, r_eq_sub, 8, fill=SUBTOTAL_FILL, font=BOLD_FONT)

# ── FP 총합계 ──
r_fp_total = r_eq_sub + 2
ws2.cell(row=r_fp_total, column=1, value='기능점수(FP) 합계').font = SUBTITLE_FONT
ws2.merge_cells(start_row=r_fp_total, start_column=1, end_row=r_fp_total, end_column=6)

# Summary table
summary_items = [
    ('ILF (내부논리파일)', f'{len(ilf_data)}개', r_ilf_sub),
    ('EIF (외부인터페이스파일)', f'{len(eif_data)}개', r_eif_sub),
    ('EI (외부입력)', f'{len(ei_data)}개', r_ei_sub),
    ('EO (외부출력)', f'{len(eo_data)}개', r_eo_sub),
    ('EQ (외부조회)', f'{len(eq_data)}개', r_eq_sub),
]
for i, (name, cnt, ref_row) in enumerate(summary_items):
    rr = r_fp_total + 1 + i
    ws2.cell(row=rr, column=2, value=name)
    ws2.cell(row=rr, column=6, value=cnt)
    ws2.cell(row=rr, column=7, value=f'=G{ref_row}')
    style_row(ws2, rr, 8, font=BOLD_FONT, aligns=aligns)

r_fp_grand = r_fp_total + 1 + len(summary_items)
ws2.cell(row=r_fp_grand, column=2, value='총 기능점수 (UFP)')
ws2.cell(row=r_fp_grand, column=7, value=f'=SUM(G{r_fp_total+1}:G{r_fp_grand-1})')
style_row(ws2, r_fp_grand, 8, fill=PatternFill('solid', fgColor='FFD966'),
          font=Font(name='맑은 고딕', bold=True, size=12))

# Store the actual row for total FP (referenced from Sheet 1)
# We need to know this row number for the formula reference
print(f"FP Total row: {r_fp_grand}")  # for debugging

set_widths(ws2, [5, 22, 42, 6, 6, 8, 8, 20])

# ═══════════════════════════════════════════════════════════════
# Sheet 3: 보정계수
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('보정계수')
ws3.sheet_properties.tabColor = '70AD47'
add_title(ws3, 'SW개발비 보정계수 산정', merge_to=6)
ws3.cell(row=2, column=1, value='근거: SW사업 대가산정 가이드(2024년 개정) 제4장 보정계수').font = SMALL_FONT
ws3.merge_cells('A2:F2')

headers3 = ['보정계수', '평가 기준', '등급', '계수 값', '산정 근거', '비고']
r = 3
# We'll build inline

corr_data = [
    ('① 규모보정계수', '', '', None, '', ''),
    ('', 'FP 규모에 따른 단계별 보정', '500FP 이상', None, '500FP 이하: 1.28, 1000: 0.98, 2000: 0.91', '「기능점수 산정」FP 합계 기준'),
    ('② 연계복잡성수준', '', '', None, '', ''),
    ('', '외부 연계 시스템 수 및 복잡성', '', None, '', ''),
    ('', '· 원격 PostgreSQL DB (시계열 동기화)', '', '', '', ''),
    ('', '· Ollama SLM API (인텐트 분류)', '', '', '', ''),
    ('', '· SNMP 네트워크 장비 (13대 폴링)', '', '', '', ''),
    ('', '· Node-RED 경보분석 엔진', '', '', '', ''),
    ('', '→ 4개 외부 시스템 연계', '높음', 1.06, '4개 이상 외부 시스템 연동', ''),
    ('③ 성능요구수준', '', '', None, '', ''),
    ('', '실시간 모니터링 (2초 응답)', '', '', '', ''),
    ('', 'SSE 스트리밍 (채팅 실시간)', '', '', '', ''),
    ('', '4계층 이상감지 (캐시 84→2초)', '', '', '', ''),
    ('', '→ 응답시간/처리량 요구 높음', '높음', 1.04, '실시간 모니터링 + SSE + 캐시 최적화', ''),
    ('④ 운영환경호환성', '', '', None, '', ''),
    ('', '웹 기반 단일 플랫폼', '보통', 1.00, '크로스브라우저 (Chrome/Edge)', ''),
    ('⑤ 보안수준', '', '', None, '', ''),
    ('', 'JWT 인증 + HTTPS + 역할 기반 접근제어', '보통', 1.00, 'NextAuth + bcrypt + RBAC', ''),
    ('종합 보정계수', '', '', None, '', ''),
]

r = 3
for i, d in enumerate(corr_data):
    rr = r + i
    for c, v in enumerate(d, 1):
        if v is not None:
            ws3.cell(row=rr, column=c, value=v)
    is_section = d[0].startswith('①') or d[0].startswith('②') or d[0].startswith('③') or d[0].startswith('④') or d[0].startswith('⑤') or d[0] == '종합 보정계수'
    fill = CATEGORY_FILL if is_section else None
    font = BOLD_FONT if is_section else NORMAL_FONT
    aligns6 = [LEFT_WRAP, LEFT_WRAP, CENTER, RIGHT, LEFT_WRAP, LEFT_WRAP]
    style_row(ws3, rr, 6, fill=fill, font=font, aligns=aligns6)
    if d[3] is not None and isinstance(d[3], (int, float)):
        ws3.cell(row=rr, column=4).number_format = '0.00'
        ws3.cell(row=rr, column=4).font = BLUE_TEXT

# 규모보정계수 수식 (FP 기반 보간 — 간이 공식)
# 500 이하: 1.28, 500-1000: 1.153-0.000354*FP, 1000-2000: 0.98-0.000085*(FP-1000), ...
# 간단히 IF 수식으로
r_scale = 3  # ① 규모보정계수 행
ws3.cell(row=r_scale+1, column=4,
         value=f"=IF('기능점수 산정'!H{r_fp_grand}<=500,1.28,"
               f"IF('기능점수 산정'!H{r_fp_grand}<=1000,1.153-0.000354*('기능점수 산정'!H{r_fp_grand}-500),"
               f"IF('기능점수 산정'!H{r_fp_grand}<=2000,0.976-0.000065*('기능점수 산정'!H{r_fp_grand}-1000),"
               f"IF('기능점수 산정'!H{r_fp_grand}<=3000,0.911-0.000045*('기능점수 산정'!H{r_fp_grand}-2000),0.866))))")
ws3.cell(row=r_scale+1, column=4).font = BLUE_TEXT
ws3.cell(row=r_scale+1, column=4).number_format = '0.000'

# 보정계수 참조 행 번호 (Sheet1 에서 참조)
# ① D4, ② D11, ③ D16, ④ D18, ⑤ D20
# 종합 = ①×②×③×④×⑤
r_corr_total = r + len(corr_data) - 1
ws3.cell(row=r_corr_total, column=4, value='=D4*D11*D16*D18*D20')
ws3.cell(row=r_corr_total, column=4).number_format = '0.0000'
style_row(ws3, r_corr_total, 6, fill=PatternFill('solid', fgColor='FFD966'),
          font=Font(name='맑은 고딕', bold=True, size=11))
ws3.cell(row=r_corr_total, column=1).alignment = LEFT_WRAP
ws3.cell(row=r_corr_total, column=4).font = Font(name='맑은 고딕', bold=True, size=11, color='0000FF')

# Sheet1 보정계수 참조 업데이트
# Row mapping: ① → row 4, ② → row 11, ③ → row 16, ④ → row 18, ⑤ → row 20, 종합 → r_corr_total
# Update Sheet1 references
ws1['G11'] = f"='보정계수'!D4"   # 규모
ws1['G12'] = f"='보정계수'!D11"  # 연계
ws1['G13'] = f"='보정계수'!D16"  # 성능
ws1['G14'] = f"='보정계수'!D18"  # 운영환경
ws1['G15'] = f"='보정계수'!D20"  # 보안
ws1['G16'] = f"='보정계수'!D{r_corr_total}"  # 종합

# Fix Sheet1 FP reference
ws1[f'E{R_FP}'] = f"='기능점수 산정'!G{r_fp_grand}"

set_widths(ws3, [18, 40, 12, 10, 38, 20])

# ═══════════════════════════════════════════════════════════════
# Sheet 4: 화면 목록
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('화면 목록')
ws4.sheet_properties.tabColor = 'ED7D31'
add_title(ws4, '화면(페이지) 목록 — 34개 화면', merge_to=7)

headers4 = ['No', '메뉴 그룹', '화면명', '경로', '주요 기능', '난이도', '비고']
r = 3
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 7)

screens = [
    ('랜딩', '랜딩 페이지', '/', '시스템 소개, 로그인 연결', '하', ''),
    ('인증', '로그인', '/login', 'NextAuth 인증, JWT', '중', ''),
    ('대시보드', '종합 현황판', '/dashboard', 'KPI 6종, 이상시설 TOP, 유량불균형, 설비장애, 데이터품질, 최근경보, 팝업 분석', '상', ''),
    ('AI 채팅', 'AI 어시스턴트', '/chat', 'SSE 스트리밍, 68인텐트, 시각화, 히스토리, 자동질문', '최상', '핵심'),
    ('모니터링', '배수지', '/monitoring/reservoir', '트렌드 차트, HH/LL, 재생, 네비게이션', '상', ''),
    ('', '가압장', '/monitoring/booster', '트렌드 차트, 알람 마크라인', '상', ''),
    ('', '감압시설', '/monitoring/pressure', '트렌드 차트', '중', ''),
    ('', '블록', '/monitoring/block', '트렌드 차트', '중', ''),
    ('', '용수 흐름', '/monitoring/flow', 'SVG 계통도, 실시간, 파티클, 미니맵, 검색, LOD, 타임라인, 트렌드', '최상', '핵심'),
    ('트렌드', '트렌드 분석', '/trend', '다중 태그 비교, 태그 브라우저, 이동평균', '상', ''),
    ('네트워크', '토폴로지', '/network', 'Force/계층형 듀얼, SNMP 포트, 노드 상세', '상', ''),
    ('알람', '알람 알림', '/alarm', '자동 스캔, 실시간 알림', '중', ''),
    ('위기대응', '경보관리', '/crisis/alarm-dashboard', '2탭 (현황 도넛/이력 필터)', '상', ''),
    ('', '경보분석', '/crisis/alarm-analysis', '경보 상세, 진단, iframe', '상', ''),
    ('', '경보이력', '/crisis/alarm-history', 'redirect', '하', ''),
    ('', '현장조치', '/crisis/task-management', '조치 관리', '중', ''),
    ('구축', '태그 마스터', '/setup/tags', '2,698건, 5종 필터, CSV', '상', ''),
    ('', '설비 관리', '/setup/equipments', '290건, 3종 필터, 자동ID, CSV', '상', ''),
    ('', '배수지', '/setup/reservoir', '3탭 폼, 설비메타 16항목, CSV', '상', ''),
    ('', '가압장', '/setup/booster', '2탭 폼, 설비메타 26항목, CSV', '상', ''),
    ('', '감압시설', '/setup/pressure', '2탭 폼, CSV', '중', ''),
    ('', '블록', '/setup/block', '2탭 폼, 블록레벨, CSV', '중', ''),
    ('', '용수 흐름 계통도', '/setup/flow-map', 'SVG 계통도, CRUD, CSV', '상', ''),
    ('', '네트워크 관리', '/setup/networks', '장비탭+연결탭, 토폴로지 미리보기', '상', ''),
    ('', '트렌드 설정', '/setup/trends', '카탈로그 CRUD', '중', ''),
    ('', '캔버스 에디터', '/setup/canvas', 'React Flow, Undo/Redo, 자동배치, 내보내기, 인과', '최상', ''),
    ('', '인과 규칙', '/setup/causal-rules', '5탭 템플릿, 커버리지, 매핑 현황', '상', ''),
    ('', '컬럼 잠금', '/setup/field-locks', '컬럼별 접근 제어', '중', ''),
    ('관리자', '사용자', '/admin/users', 'CRUD, 역할 배정', '중', ''),
    ('', '메뉴', '/admin/menus', '트리 CRUD', '중', ''),
    ('', 'FAQ', '/admin/faq', 'CRUD', '중', ''),
    ('', '프롬프트', '/admin/prompts', '인텐트별 편집', '중', ''),
    ('', '시설 파일', '/admin/facility-files', '이미지 업로드', '중', ''),
    ('', '사이트 설정', '/admin/site-settings', '시스템 설정', '하', ''),
]

for i, s in enumerate(screens):
    r = 4 + i
    ws4.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(s, 2):
        ws4.cell(row=r, column=c, value=v)
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, CENTER, LEFT_WRAP]
    style_row(ws4, r, 7, fill=CATEGORY_FILL if s[0] else None, font=BOLD_FONT if s[0] else NORMAL_FONT, aligns=aligns)

set_widths(ws4, [5, 10, 18, 28, 42, 8, 12])

# ═══════════════════════════════════════════════════════════════
# Sheet 5: DB 테이블 목록
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('DB 테이블 목록')
ws5.sheet_properties.tabColor = 'BF8F00'
add_title(ws5, 'DB 설계 목록 (PostgreSQL 16 + TimescaleDB)', merge_to=7)

headers5 = ['No', '카테고리', '테이블/뷰명', '유형', '설명', '주요 컬럼', '비고']
r = 3
for c, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, 7)

tables = [
    ('인증', 'tb_auth', '테이블', '권한 그룹', 'auth_idn, auth_name, auth_level', ''),
    ('', 'tb_user', '테이블', '사용자 계정 (bcrypt)', 'user_id, pw, pw_migrated, auth_idn, region', ''),
    ('', 'tb_user_session', '테이블', '로그인 세션', 'session_id, user_id, expires_at', ''),
    ('', 'tb_access_log', '테이블', 'API 접근 감사 로그', 'log_id, user_id, api_path, timestamp', ''),
    ('메뉴', 'tb_menu', '테이블', '메뉴 트리 (자기참조)', 'menu_idn, pmenu_idn, menu_name, app_path', ''),
    ('', 'tb_auth_menu', '테이블', '권한↔메뉴 매핑', 'auth_idn, menu_idn, read_yn, write_yn', ''),
    ('', 'tb_menu_api', '테이블', '메뉴↔API 매핑', 'menu_idn, api_path, method', ''),
    ('코드', 'tb_grp_code', '테이블', '그룹 코드 마스터', 'grp_code, grp_name', ''),
    ('', 'tb_comm_code', '테이블', '상세 코드', 'grp_code, comm_code, comm_name', ''),
    ('채팅', 'tb_ai_chat_ask_group', '테이블', '채팅 세션', 'group_id, user_id, title', ''),
    ('', 'tb_ai_chat_ask', '테이블', '사용자 질문', 'ask_id, group_id, question, ask_at', ''),
    ('', 'tb_ai_chat_bot', '테이블', '봇 응답 (JSONB)', 'bot_id, ask_id, answer, visual_data', '차트 재렌더링'),
    ('', 'tb_ai_chat_ask_image', '테이블', '질문 이미지', 'image_id, ask_id, file_path', ''),
    ('', 'tb_ai_chat_bot_image', '테이블', '응답 이미지', 'image_id, bot_id, file_path', ''),
    ('', 'tb_ai_chat_faq', '테이블', 'FAQ', 'faq_id, question, category', ''),
    ('시설', 'tb_service_reservoir_info', '테이블', '배수지 정보', 'sitename, region, capacity_m3', 'JSONB'),
    ('', 'tb_service_reservoir_status', '테이블', '배수지 상태', 'sitename, hh, ll, target_level, meta', '16항목'),
    ('', 'tb_service_booster_station_info', '테이블', '가압장 정보', 'sitename, region, facilitytype', ''),
    ('', 'tb_service_booster_station_status', '테이블', '가압장 상태', 'sitename, hh, ll, meta', '26항목'),
    ('', 'tb_pressure_reducing_facility_info', '테이블', '감압시설 정보', 'sitename, region', ''),
    ('', 'tb_pressure_reducing_facility_status', '테이블', '감압시설 상태', 'sitename, hh, ll, meta', ''),
    ('', 'tb_block_info', '테이블', '블록 정보', 'sitename, block_level', ''),
    ('', 'tb_block_status', '테이블', '블록 상태', 'sitename, target_flow, meta', ''),
    ('설비', 'tb_equipment_info', '테이블', '설비 마스터', 'equipment_id, sitename, equipment_type, ip', '290건'),
    ('', 'tb_equipment_status', '테이블', '설비 상태', 'equipment_id, is_alive, rtt_ms', ''),
    ('', 'tb_equipment_alarm_report', '테이블', '경보 리포트', 'alarm_id, equipment_id, alarm_type, start', ''),
    ('', 'tb_equipment_tag_map', '테이블', '설비↔태그 매핑', 'equipment_id, tagsn, mapped_by', '3,375건'),
    ('네트워크', 'tb_network_link', '테이블', '네트워크 연결', 'source_id, target_id, protocol', ''),
    ('', 'tb_facility_flow_map', '테이블', '용수 계통도', 'source_sitename, target_sitename', '95엣지'),
    ('태그', 'tb_tag_info', '테이블', '태그 마스터', 'tagsn, sitename, facilitytype, datainfo, unit', '2,698건'),
    ('', 'tb_tag_raw_data', '하이퍼테이블', '시계열 데이터', 'tagsn, timestamp, value, quality', '2.3M+행'),
    ('', 'tb_tag_data_group', '테이블', '태그 그룹 (21개)', 'group_code, parent_code, keywords', '계층'),
    ('', 'tb_tag_group_map', '테이블', '태그↔그룹 매핑', 'tagsn, group_code, matched_keyword', '93%자동'),
    ('', 'tb_trend_catalog', '테이블', '트렌드 카탈로그', 'catalog_id, name, items', ''),
    ('', 'tb_alarm_log', '테이블', '알람 이력', 'tagsn, alarm_type, start_time, resolved', ''),
    ('모니터링', 'tb_monitoring_catalog', '테이블', '모니터링 카탈로그', 'catalog_id, catalog_name, items', ''),
    ('', 'tb_admin_site_settings', '테이블', '사이트 설정', 'setting_key, setting_value', ''),
    ('', 'tb_site_anomaly_profile', '테이블', '현장 프로파일', 'sitename, group, p95, p05', '일 1회'),
    ('캔버스', 'tb_canvas_node_position', '테이블', '노드 위치', 'sitename, facilitytype, x, y', ''),
    ('', 'tb_causal_chain_override', '테이블', '인과 오버라이드', 'sitename, facilitytype, zone, chain', 'JSONB'),
    ('', 'tb_snmp_port_status', '테이블', 'SNMP 포트', 'equipment_id, port_index, status, speed', ''),
    ('관리', 'tb_prompt_template', '테이블', '프롬프트 템플릿', 'intent, template', ''),
    ('', 'tb_prompt_column', '테이블', '프롬프트 컬럼', 'intent, column_name, column_type', ''),
    ('', 'tb_file_storage', '테이블', '파일 메타데이터', 'file_id, file_path, mime_type', ''),
    ('', 'tb_file_history', '테이블', '파일 버전 이력', 'history_id, file_id, version', ''),
    ('', 'tb_facility_file', '테이블', '시설↔파일 매핑', 'facility_file_id, sitename, file_id', ''),
    ('', 'tb_field_lock', '테이블', '컬럼 잠금', 'table_name, column_name, locked_by', ''),
    ('뷰', 'v_reservoir_info_status', '뷰', '배수지 통합뷰', '', ''),
    ('', 'v_booster_station_info_status', '뷰', '가압장 통합뷰', '', ''),
    ('', 'v_pressure_reducing_facility_info_status', '뷰', '감압시설 통합뷰', '', ''),
    ('', 'v_block_info_status', '뷰', '블록 통합뷰', '', ''),
    ('', 'v_ongoing_alarm', '뷰', '진행 중 알람', '', ''),
    ('', 'cagg_5min_raw_stats_ai', '연속집계', '5분 통계', '', 'TimescaleDB'),
    ('', 'mv_tag_daily_status', '물리화뷰', '일별 통계', '', ''),
]

for i, t in enumerate(tables):
    r = 4 + i
    ws5.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(t, 2):
        ws5.cell(row=r, column=c, value=v)
    type_cell = ws5.cell(row=r, column=4)
    if t[2] == '하이퍼테이블': type_cell.fill = PatternFill('solid', fgColor='E8D5F5')
    elif t[2] in ('뷰', '연속집계', '물리화뷰'): type_cell.fill = PatternFill('solid', fgColor='D5E8F5')
    aligns = [CENTER, LEFT_WRAP, LEFT_WRAP, CENTER, LEFT_WRAP, LEFT_WRAP, LEFT_WRAP]
    style_row(ws5, r, 7, fill=CATEGORY_FILL if t[0] else None, font=BOLD_FONT if t[0] else NORMAL_FONT, aligns=aligns)

r_db_total = 4 + len(tables)
tbl_cnt = sum(1 for t in tables if t[2] == '테이블')
etc_cnt = len(tables) - tbl_cnt
ws5.cell(row=r_db_total, column=1, value=f'합계: 테이블 {tbl_cnt}개 + 하이퍼테이블 1개 + 뷰/집계 {etc_cnt-1}개 = {len(tables)}개')
ws5.merge_cells(start_row=r_db_total, start_column=1, end_row=r_db_total, end_column=7)
style_row(ws5, r_db_total, 7, fill=TOTAL_FILL, font=BOLD_FONT)

set_widths(ws5, [5, 8, 38, 10, 22, 38, 12])

# ═══════════════════════════════════════════════════════════════
# Sheet 6: 코드 규모 분석
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('코드 규모 분석')
ws6.sheet_properties.tabColor = 'FFC000'
add_title(ws6, '코드 규모 분석 (LOC 기반)', merge_to=6)

headers6 = ['구분', '분류', '파일 수', 'LOC', '평균 LOC', '설명']
r = 3
for c, h in enumerate(headers6, 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, 6)

loc_data = [
    ('프론트엔드', '컴포넌트 (React)', 163, 28319, None, 'UI/CRUD/차트/캔버스'),
    ('', '페이지 (page.tsx)', 34, 7466, None, '라우트별 페이지'),
    ('', '차트 옵션·설정', 15, 3255, None, 'ECharts 옵션'),
    ('', 'API 클라이언트', 27, 2548, None, 'fetch 래퍼'),
    ('', '타입 정의', 19, 1891, None, 'TypeScript'),
    ('', '상태 관리 (Zustand)', 17, 1712, None, '17개 스토어'),
    ('', '커스텀 훅', 11, 1537, None, '재사용 로직'),
    ('', '유틸리티·기타', 50, 4269, None, ''),
    ('프론트엔드 소계', '', 336, 50997, None, ''),
    ('백엔드', 'ai_server.py', 1, 14846, None, '106+ API'),
    ('', '이상감지 모듈', 3, 2926, None, '4계층 탐지'),
    ('', '인텐트 분류', 4, 1951, None, '3단계'),
    ('', '물 수지 검증', 1, 443, None, '유량적분'),
    ('', 'SNMP 폴링', 1, 414, None, '13대'),
    ('', 'DB 동기화', 2, 524, None, ''),
    ('', '기타 모듈', 12, 2071, None, ''),
    ('백엔드 소계', '', 24, 23175, None, ''),
    ('전체 합계', '', 360, 74172, None, ''),
]

for i, d in enumerate(loc_data):
    r = 4 + i
    for c, v in enumerate(d, 1):
        ws6.cell(row=r, column=c, value=v)
    if d[4] is None and d[2] and d[3]:
        ws6.cell(row=r, column=5, value=f'=IF(C{r}>0,ROUND(D{r}/C{r},0),"")')
    is_total = '소계' in str(d[0]) or '합계' in str(d[0])
    aligns = [LEFT_WRAP, LEFT_WRAP, RIGHT, RIGHT, RIGHT, LEFT_WRAP]
    style_row(ws6, r, 6, fill=TOTAL_FILL if is_total else (CATEGORY_FILL if d[0] and not is_total else None),
              font=BOLD_FONT if (d[0] or is_total) else NORMAL_FONT, aligns=aligns)
    ws6.cell(row=r, column=3).number_format = NUM_FMT
    ws6.cell(row=r, column=4).number_format = NUM_FMT
    ws6.cell(row=r, column=5).number_format = NUM_FMT

set_widths(ws6, [16, 22, 10, 10, 10, 25])

# ── 인쇄 설정 ──
for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

# ── FP 합계 행을 Sheet2 Column H에도 복사 (Sheet1 참조용) ──
# Sheet1에서 '기능점수 산정'!H{r_fp_grand}를 참조하므로 G→H 복사
ws2.cell(row=r_fp_grand, column=8, value=f'=G{r_fp_grand}')

OUTPUT = 'D:/web/SLM_대가산정표_FP.xlsx'
wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
print(f'FP Total Row: {r_fp_grand} (Sheet "기능점수 산정", Column G/H)')
print(f'Correction Total Row: {r_corr_total} (Sheet "보정계수", Column D)')
