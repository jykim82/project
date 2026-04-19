"""
SW사업 대가산정 가이드 (2024, KOSA) — 단일 FP유형 (간이법 평균복잡도)
참고: docs/2. SW사업 구현단계_관망관리_시설물관리 합본_단일FP유형.xlsx
"""
import sys, io, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 스타일 ──
HDR_FONT = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='2F5496')
SUB_HDR_FILL = PatternFill('solid', fgColor='4472C4')
CAT_FILL = PatternFill('solid', fgColor='D6E4F0')
TOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')
SUB_FILL = PatternFill('solid', fgColor='E2EFDA')
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=13)
SEC_FONT = Font(name='맑은 고딕', bold=True, size=11, color='2F5496')
NF = Font(name='맑은 고딕', size=10)
BF = Font(name='맑은 고딕', bold=True, size=10)
BLUE = Font(name='맑은 고딕', size=10, color='0000FF')
SMALL = Font(name='맑은 고딕', size=9, color='666666')
thin = Side(style='thin', color='B4C6E7')
BD = Border(left=thin, right=thin, top=thin, bottom=thin)
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)
R = Alignment(horizontal='right', vertical='center')
NUM = '#,##0'
PCT = '0.00'

def hdr(ws, row, ncol):
    for c in range(1, ncol+1):
        cl = ws.cell(row=row, column=c)
        cl.font = HDR_FONT; cl.fill = HDR_FILL; cl.alignment = C; cl.border = BD

def srow(ws, row, ncol, fill=None, font=None, als=None):
    for c in range(1, ncol+1):
        cl = ws.cell(row=row, column=c)
        cl.font = font or NF; cl.border = BD
        if fill: cl.fill = fill
        if als and c <= len(als): cl.alignment = als[c-1]
        else: cl.alignment = C

def title(ws, text, row=1, mc=8):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)

def widths(ws, ww):
    for c, w in enumerate(ww, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

# ── 단일 FP유형 가중치 (간이법 평균복잡도) ──
FP_W = {'ILF': 7.5, 'EIF': 5.4, 'EI': 4.0, 'EO': 5.2, 'EQ': 3.9}
FP_UNIT_PRICE = 605784  # 2025년 고시 단가 (원/FP)

# ═══════════════════════════════════════════════════════════════
# Sheet 1: FP산정(간이법)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'FP산정(간이법)'
ws1.sheet_properties.tabColor = '2F5496'

title(ws1, '기능점수 산정 (간이법 평균복잡도 적용)', mc=9)

# Headers (참고 문서와 동일 구조)
r = 3
for c, h in enumerate(['기능명', '', '', '', '', '데이터 및 트랜잭션 기능', '', 'FP 산출', '사용자관점 기능'], 1):
    ws1.cell(row=r, column=c, value=h)
hdr(ws1, r, 9)
ws1.merge_cells('A3:A3')
ws1.merge_cells('F3:G3')

r = 4
sub_headers = ['①어플리케이션명', '②세부 업무명', '③단위프로세스명', '단위프로세스 설명', '기능 상세 설명', '', '④FP유형', '⑤가중치', '']
for c, h in enumerate(sub_headers, 1):
    ws1.cell(row=r, column=c, value=h)
hdr(ws1, r, 9)

# ── 기능 목록 ──
# (app, sub, process, desc, detail, fp_type, user_yn)
features = [
    # === 1. 종합 현황판 ===
    (' < SLM 스마트 관망관리 시스템 >', '1. 종합 현황판', '실시간 상황 감시', '종합 현황판', '시설유형별 KPI 6종, 이상시설 TOP, 유량불균형, 설비장애, 데이터품질, 최근경보를 시각화하여 표출', 'ILF', 'Y'),
    ('', '', '', '실시간 종합상황판', '캐시 집계 기반 종합 현황, 이상감지·교차검증·물수지·설비장애를 통합 표시', 'EO', ''),
    ('', '', '', '현황판 클릭 분석', '항목 클릭 → QuickAnalysisDialog 팝업 분석, 경보 상세 팝업', 'EO', ''),
    ('', '', '', '현황판 설정', '대시보드 자료 등록, 수정, 삭제, 조회 설정관리', 'EI', ''),
    ('', '', '실시간 이벤트 감시', '계측 데이터 이상 감시', '4계층 이상감지(인과+홀딩+Z-Score+물수지)로 이상 발생 시 알람 발생', 'ILF', 'Y'),
    ('', '', '', '이상감지 전체 스캔', '전체 센서 Z-Score + 교차검증 + 물수지 + 데이터품질 + 설비장애 통합 스캔', 'EO', ''),
    ('', '', '', '이상감지 시설 상세', '인과관계 검증 + 교차시설 검증 + 전파 추적 + 시설 내부 검증 + SLM 해석', 'EO', ''),
    ('', '', '', '이벤트 조치 안내', '실시간 알림으로 표출된 이상에 대한 인과관계 판정 및 조치 안내', 'EO', ''),
    ('', '', '종합현황', '관망운영 종합일보', '시설유형별 실시간 상태, 이상 현황, 경보 현황 종합 보고', 'ILF', 'Y'),
    ('', '', '', '교차시설 검증 현황', '상류→하류 유량/압력 흐름 일관성 자동 검증 결과 표시', 'EO', ''),
    ('', '', '', '물 수지 현황', '상류 유출유량 vs 하류 유입유량 합계 비교, 불균형 등급 표시', 'EO', ''),
    ('', '', '', '이벤트 현황', '이상감지, 교차검증, 설비장애 등 이벤트 발생현황 보고', 'EO', ''),
    ('', '', '', '이벤트 통계분석', '시간별, 시설유형별, 이벤트 종류별 발생·조치 현황 분류', 'EIF', ''),
    ('', '', '데이터 품질관리', '오·결측 데이터 관리', '센서무응답, 데이터홀딩, 데이터없음, 데이터부족 4종 감지 및 표시', 'EO', 'Y'),
    ('', '', '', '홀딩 감지', '5분 버킷 flat + 시간 단위 집계, 기준선에서 홀딩 시간 제외', 'ILF', ''),
    ('', '', '', '데이터 수집관리', '태그 관리, 태그 그룹(21종) 자동분류, 수집데이터 상태 모니터링', 'EIF', ''),

    # === 2. AI 채팅 어시스턴트 ===
    ('', '2. AI 채팅 어시스턴트', 'AI 인텐트 분류', 'AI 채팅 질의 (동기)', '3단계 분류(키워드→벡터→SLM) → SQL 실행 → 응답 생성, 68인텐트', 'EI', 'Y'),
    ('', '', '', 'AI 채팅 질의 (SSE)', 'SSE 스트리밍 (token→answer_complete→visual→recommend→done)', 'EO', ''),
    ('', '', '', '인텐트 보정', '오분류 시 force_intent 파라미터로 사후 보정, 대체 후보 3개 표시', 'EI', ''),
    ('', '', '', '인텐트 벡터 검색', 'snowflake-arctic-embed2 (1024dim) 임베딩, cosine 유사도, NPY 캐시', 'EIF', ''),
    ('', '', '채팅 시각화', '시각화 렌더링', '그래프(트렌드/산점도), 테이블, 다이어그램, 문서 5종 시각화 자동 라우팅', 'EO', 'Y'),
    ('', '', '', '채팅 히스토리', '대화 이력 저장, visual_data JSONB로 차트 재렌더링', 'ILF', ''),
    ('', '', '', '추천 질의', '인텐트별 후속 추천질의 3건, 현장명 자동 보강', 'EO', ''),
    ('', '', '', '자동 질문 제출', '대시보드/팝업에서 ?q= 파라미터로 채팅 자동 질문', 'EI', ''),

    # === 3. 모니터링 ===
    ('', '3. 실시간 모니터링', '시설별 트렌드 모니터링', '배수지/가압장/감압/블록 모니터링', '트렌드 차트, HH/LL 가이드선, 재생, 시간 네비게이션, 이동평균', 'ILF', 'Y'),
    ('', '', '', '트렌드 차트 재생', '80ms 프레임 setOption 방식, 확장 모드 연동', 'EO', ''),
    ('', '', '', '시간 네비게이션', '좌/우 시간 이동, 이전 데이터 로드, viewOffset 관리', 'EO', ''),
    ('', '', '', '차트 통계 바', '↑max ≈avg ↓min 통계, 아날로그/디지털 혼합 대응', 'EO', ''),
    ('', '', '', '모니터링 임계치 설정', 'HH/LL 알람 설정 듀얼모드 (상수/태그참조 선택)', 'EI', ''),
    ('', '', '용수 흐름 실시간 모니터링', '용수 흐름 계통도', 'SVG 계통도, 유량 비례 엣지, 파티클 애니메이션, 미니맵, 검색, LOD', 'ILF', 'Y'),
    ('', '', '', '실시간 유량/수위/압력', '노드별 Q(유량)/H(수위)/P(압력)/T(공급시간) 수치 오버레이, 기준선 비율', 'EO', ''),
    ('', '', '', '교차검증 이상 표시', '상류 유형별(배수지 80%/가압장 30%) 불일치 감지, 노드 펄스 링', 'EO', ''),
    ('', '', '', '물수지 불균형 표시', '엣지 색상(경고=빨강/주의=주황/관심=노랑) + % 뱃지', 'EO', ''),
    ('', '', '', '설비 장애 표출', '4종 뱃지(NET/COM/FLT/PWR), 24시설 감지, 알람 경고 링', 'EO', ''),
    ('', '', '', '타임라인 재생', '과거 24h 5분 간격 슬라이더 재생, 실시간↔타임라인 전환', 'EO', ''),
    ('', '', '', '배수지 구역 선택', '다중 수위 태그(1지/2지, 공업/생활) 탭 전환', 'EO', ''),
    ('', '', '', '트렌드 패널', '노드 클릭 → 하단 fixed 패널 24h 스파크라인(Canvas)', 'EO', ''),
    ('', '', '', 'KPI 카드 필터', '유량없음/교차이상/불균형/통신이상/알람/설비장애 7종 클릭 필터', 'EI', ''),
    ('', '', '경보 경보 관리', '경보관리 (현황/이력)', '2탭 구조(현황 도넛+카테고리+테이블 / 이력 필터+확인)', 'ILF', 'Y'),
    ('', '', '', '경보분석', '경보분석 상세, 진단 메시지 iframe, 경보 확인 처리', 'EO', ''),
    ('', '', '', '경보 이상지점 조회', '통신/수위/압력/펌프/밸브/유량/전원 × 시설유형 × HH/LL/FAULT', 'EO', ''),
    ('', '', '', '알람 알림', '자동 스캔, 실시간 알림 벨, 진행중 건수', 'EO', ''),

    # === 4. 트렌드 분석 ===
    ('', '4. 트렌드 분석', '트렌드 태그 조합 비교', '트렌드 분석 페이지', '최대 15태그 자유조합, 5종 필터+페이징 태그 브라우저', 'ILF', 'Y'),
    ('', '', '', '시계열 데이터 조회', 'TimescaleDB 청크 직접쿼리, time_bucket 집계, Python JOIN', 'EO', ''),
    ('', '', '야간최소유량', '야간최소유량 표', '60분 이동평균, 청크 직접쿼리(31.9→7.5초)', 'EO', 'Y'),
    ('', '', '', '야간최소유량 표준편차', '400일 1회 조회 + Python 통계(avg/stddev/신뢰구간/초과량)', 'EO', ''),
    ('', '', '', '다중 시설 표준편차 뷰', '전체 소블록 표준편차 분석, KPI 4종 + 비교 바 차트 + 미니 카드', 'EO', ''),
    ('', '', '결측분석', '결측분석 표', 'SQL 분단위 집계(15.5→2.8초), 홀딩 감지, 일별 통계', 'EO', 'Y'),
    ('', '', '카탈로그 트렌드', '카탈로그 트렌드 표', '2단계 청크 직접쿼리, 시설유형×데이터키워드', 'EO', ''),

    # === 5. 이상감지 엔진 ===
    ('', '5. 이상감지 엔진', '인과관계 Rule 엔진', '시설유형별 인과 체인', '5개 시설유형 물리법칙 기반 인과 템플릿, 구역(1지/2지) 분리', 'ILF', 'Y'),
    ('', '', '', '인과 검증', 'verify_causal_context(): 5가지 불일치 패턴 판정, 형제 그룹 매칭', 'EO', ''),
    ('', '', '', '시설 내부 인과 검증', '펌프/밸브/수위/압력 6개 규칙(가압장3+배수지3+감압1)', 'EO', ''),
    ('', '', '', '교차상관 시간지연 추정', 'numpy 교차상관, Z-score 정규화, positive lag', 'EO', ''),
    ('', '', '', 'SLM 자연어 해석', 'Phi-4-mini 인과 판정 결과 → 3문장 자연어 해석', 'EO', ''),
    ('', '', '교차시설 검증', '상류→하류 유량 일관성', 'BFS 기반 하류 전파 + 상류 근원지 역추적 (max_depth=3)', 'EO', 'Y'),
    ('', '', '', '다중 홉 전파 추적', 'trace_propagation_forward/upstream_root_cause, 근원지 confidence', 'EO', ''),
    ('', '', '물 수지 검증', '유량 적분 Mass Balance', '사다리꼴 적분, 5등급(정상/관심/주의/경고), 24h 롤링', 'EO', 'Y'),
    ('', '', '현장 프로파일링', 'A/B/C/D 그룹 분류', '일 1회 유출유량/알람빈도/P95/P05, GROUP_THRESHOLDS', 'ILF', ''),
    ('', '', '', 'Isolation Forest', '백그라운드 학습(24h 주기), per-facility 322개 모델', 'EIF', ''),
    ('', '', '설비 장애 역추적', '설비 통신 끊김/장애 감지', '네트워크 is_alive + DI 장애 3종 → tb_equipment_tag_map 역추적', 'EO', ''),
    ('', '', '', '종합 판정', 'compute_verdict 5단계: 복합이상/이상/교차이상/주의/정상', 'EO', ''),

    # === 6. 네트워크 관리 ===
    ('', '6. 네트워크 관리', '토폴로지 시각화', 'Force/계층형 듀얼 뷰', 'ECharts Force + SVG 3컬럼 적응형, 노드 상세 패널', 'ILF', 'Y'),
    ('', '', '', 'SNMP 스위치 포트', 'SVG 포트 정면 다이어그램, 13대 24포트, 3분 주기 폴링', 'EO', ''),
    ('', '', '', '네트워크 연결 플로우', '행 더블클릭 → BFS 서브그래프 + SVG 미니 트리 팝업', 'EO', ''),
    ('', '', '', '네트워크 CRUD', '장비/연결 CRUD, CSV 업/다운, 토폴로지 미리보기', 'EI', ''),

    # === 7. 구축(Setup) ===
    ('', '7. 기초정보 구축', '태그 마스터 관리', '태그 마스터', '2,698건, 5종 필터+페이징, 태그추가/엑셀업로드/CSV다운로드', 'ILF', 'Y'),
    ('', '', '', '태그 그룹 분류', '21그룹 계층구조, longest-keyword-first 93% 자동분류', 'EI', ''),
    ('', '', '설비 관리', '설비 CRUD', '290건, 3종 필터, 접두사→자동ID, CASCADE 삭제 체크', 'ILF', 'Y'),
    ('', '', '', '설비↔태그 자동 매핑', '4종 규칙(가압펌프/PLC/LTE/유량계) 3,375건 자동생성', 'EI', ''),
    ('', '', '배수지 관리', '배수지 CRUD', '3탭 폼(기본/구역/운영), 설비메타 16항목, CSV 27컬럼', 'ILF', 'Y'),
    ('', '', '가압장 관리', '가압장 CRUD', '2탭 폼(기본/운영), 설비메타 26항목, CSV 18컬럼', 'ILF', 'Y'),
    ('', '', '감압시설 관리', '감압시설 CRUD', '2탭 폼, 설비메타 5항목, CSV 9컬럼', 'ILF', ''),
    ('', '', '블록 관리', '블록 CRUD', '2탭 폼, 블록레벨 필터, CSV 11컬럼', 'ILF', ''),
    ('', '', '용수 흐름 관리', '계통도 CRUD', 'SVG 계통도, 95엣지, CSV 입출력', 'ILF', 'Y'),
    ('', '', 'CSV 통합', 'CSV 다운로드/업로드 (8종)', '빈 템플릿 지원, CsvUploadDialog 범용 컴포넌트', 'EI', ''),
    ('', '', '캔버스 에디터', '캔버스 시설 토폴로지', 'React Flow, 78노드+76엣지, DnD, Ctrl+S 자동저장', 'ILF', 'Y'),
    ('', '', '', '자동 배치', 'BFS depth + Sugiyama median 교차최소화(8라운드)', 'EO', ''),
    ('', '', '', 'Undo/Redo + 내보내기', 'Ctrl+Z/Y, MAX=50 스냅샷, PNG/SVG 내보내기', 'EI', ''),
    ('', '', '', '속성 패널 4탭', '정보/설비/태그/인과 4탭, lazy load, CRUD 연동', 'EO', ''),
    ('', '', '인과 규칙 관리', '인과 규칙 UI', '5탭 템플릿, 시설 커버리지, 태그 매핑 현황, 프로그레스 바', 'EQ', 'Y'),
    ('', '', '', '인과 오버라이드 CRUD', 'tb_causal_chain_override 저장/삭제, 구역 선택, 자동 추정', 'EI', ''),
    ('', '', '모니터링 설정', '모니터링 카탈로그', '카탈로그 CRUD, 사이트 필터, 참조 데이터', 'EI', ''),
    ('', '', '컬럼 잠금', '컬럼별 접근 제어', '필드 단위 잠금/해제', 'EI', ''),

    # === 8. 관리자 ===
    ('', '8. 시스템 관리', '사용자 관리', '사용자/역할 관리', 'CRUD, 역할 배정, bcrypt 비밀번호', 'ILF', 'Y'),
    ('', '', '메뉴 관리', '메뉴 트리 관리', '자기참조 FK, 트리 CRUD', 'ILF', ''),
    ('', '', 'FAQ 관리', 'FAQ CRUD', '추천 질문 관리', 'EI', ''),
    ('', '', '프롬프트 관리', '인텐트별 프롬프트', '프롬프트 템플릿 편집', 'EI', ''),
    ('', '', '시설 파일 관리', '이미지 업로드', 'multipart 업로드, 파일 관리', 'EI', ''),
    ('', '', '사이트 설정', '시스템 설정', '사이트 레벨 설정', 'EI', ''),

    # === 9. 공통/기반 ===
    ('', '9. 공통 기반', '인증/세션', '로그인 (NextAuth JWT)', 'CredentialsProvider, bcrypt, JWT 발급/갱신', 'EI', 'Y'),
    ('', '', '', '세션 관리', '세션 생성/만료/정리, 60초 주기 클린업', 'ILF', ''),
    ('', '', '자동완성', '자동완성 후보', '현장명/시설유형/데이터항목/블록유형/질의 템플릿', 'EQ', ''),
    ('', '', '한글 퍼지 매칭', '한글 오타 보정', '자모 분해, fuzzy stopwords, 현장명 매칭', 'EIF', ''),
    ('', '', 'DB 동기화', '원격→로컬 동기화', 'db_sync.py 증분 동기화, 시계열 + 구조 테이블', 'EIF', ''),
    ('', '', '서버 상태', '헬스체크 / 모델 관리', 'Ollama 상태, DB 연결, 모델 선택', 'EQ', ''),
]

r = 5
fp_counts = {'ILF': 0, 'EIF': 0, 'EI': 0, 'EO': 0, 'EQ': 0}

for i, f in enumerate(features):
    rr = r + i
    app, sub, proc, desc, detail, fp_type, user_yn = f
    ws1.cell(row=rr, column=1, value=app)
    ws1.cell(row=rr, column=2, value=sub)
    ws1.cell(row=rr, column=3, value=proc)
    ws1.cell(row=rr, column=4, value=desc)
    ws1.cell(row=rr, column=5, value=detail)
    ws1.cell(row=rr, column=7, value=fp_type)
    ws1.cell(row=rr, column=8, value=FP_W[fp_type])
    ws1.cell(row=rr, column=9, value=user_yn)

    fp_counts[fp_type] += 1

    als = [L, L, L, L, L, L, C, R, C]
    fill = CAT_FILL if app else None
    font = BF if app or sub else NF
    srow(ws1, rr, 9, fill=fill, font=font, als=als)

r_end = r + len(features)

# 합계
ws1.cell(row=r_end+1, column=1, value='합 계').font = BF
ws1.merge_cells(start_row=r_end+1, start_column=1, end_row=r_end+1, end_column=7)
ws1.cell(row=r_end+1, column=8, value=f'=SUM(H{r}:H{r_end-1})')
ws1.cell(row=r_end+1, column=8).number_format = '0.0'
srow(ws1, r_end+1, 9, fill=TOTAL_FILL, font=Font(name='맑은 고딕', bold=True, size=11))
ws1.cell(row=r_end+1, column=1).alignment = L

FP_TOTAL_ROW = r_end + 1  # Sheet2에서 참조

widths(ws1, [22, 18, 20, 22, 55, 3, 7, 8, 5])

# ═══════════════════════════════════════════════════════════════
# Sheet 2: SW개발비 산정
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('SW개발비 산정')
ws2.sheet_properties.tabColor = '70AD47'

ws2.cell(row=1, column=2, value='"SLM 스마트 관망관리 시스템" 소프트웨어 개발비 산정').font = TITLE_FONT
ws2.merge_cells('B1:K1')
ws2.cell(row=2, column=8, value='(소프트웨어사업 대가산정 가이드, 2024. KOSA)').font = SMALL
ws2.merge_cells('H2:L2')

# ── 개발원가 산정 ──
r = 4
ws2.cell(row=r, column=2, value='○ 개발원가 산정').font = SEC_FONT
ws2.cell(row=r, column=12, value='(단위 : 원)').font = SMALL
r = 5
for c, h in enumerate(['총기능점수', '', '', '기능점수당 단가', '', '보정계수', '', '', '', '', '개발원가'], 2):
    ws2.cell(row=r, column=c, value=h)
hdr(ws2, r, 12)
ws2.merge_cells('B5:C5')
ws2.merge_cells('D5:E5')
ws2.merge_cells('F5:J5')
r = 6
for c, h in enumerate(['', '', '', '', '', '규모', '연계복잡성', '성능', '운영환경', '보안성', ''], 2):
    ws2.cell(row=r, column=c, value=h)
hdr(ws2, r, 12)

r = 7
# 총기능점수 = FP산정 시트 합계
ws2.cell(row=r, column=2, value=f"='FP산정(간이법)'!H{FP_TOTAL_ROW}")
ws2.cell(row=r, column=2).font = BLUE
ws2.cell(row=r, column=2).number_format = '0.0'
# 단가
ws2.cell(row=r, column=4, value=FP_UNIT_PRICE)
ws2.cell(row=r, column=4).number_format = NUM
# 보정계수 (하단에서 산정)
ws2.cell(row=r, column=6, value='=K21')  # 규모
ws2.cell(row=r, column=6).font = BLUE
ws2.cell(row=r, column=7, value='=K22')  # 연계
ws2.cell(row=r, column=7).font = BLUE
ws2.cell(row=r, column=8, value='=K23')  # 성능
ws2.cell(row=r, column=8).font = BLUE
ws2.cell(row=r, column=9, value='=K24')  # 운영환경
ws2.cell(row=r, column=9).font = BLUE
ws2.cell(row=r, column=10, value='=K25')  # 보안
ws2.cell(row=r, column=10).font = BLUE
# 개발원가 = FP × 단가 × 보정계수 곱
ws2.cell(row=r, column=11, value='=B7*D7*F7*G7*H7*I7*J7')
ws2.cell(row=r, column=11).number_format = NUM
for c in range(2, 12):
    ws2.cell(row=r, column=c).border = BD
    ws2.cell(row=r, column=c).alignment = R if c not in [6,7,8,9,10] else C

r = 8
ws2.cell(row=r, column=2, value='합계(보정 후 개발원가)').font = BF
ws2.merge_cells('B8:J8')
ws2.cell(row=r, column=11, value='=K7')
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, fill=SUB_FILL, font=BF)

r = 9
ws2.cell(row=r, column=2, value='이윤').font = BF
ws2.cell(row=r, column=10, value=0.2)
ws2.cell(row=r, column=10).number_format = '0%'
ws2.cell(row=r, column=11, value='=K8*J9')
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, font=NF)

r = 10
ws2.cell(row=r, column=2, value='직접경비').font = BF
ws2.cell(row=r, column=11, value=0)
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, font=NF)

r = 11
ws2.cell(row=r, column=2, value='소프트웨어 개발비(부가세 별도)').font = Font(name='맑은 고딕', bold=True, size=11)
ws2.merge_cells('B11:J11')
ws2.cell(row=r, column=11, value='=K8+K9+K10')
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, fill=SUB_FILL, font=BF)

r = 12
ws2.cell(row=r, column=2, value='부가가치세 (10%)').font = BF
ws2.cell(row=r, column=11, value='=K11*0.1')
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, font=NF)

r = 13
ws2.cell(row=r, column=2, value='총 사업비 (VAT 포함)').font = Font(name='맑은 고딕', bold=True, size=12)
ws2.merge_cells('B13:J13')
ws2.cell(row=r, column=11, value='=K11+K12')
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, fill=PatternFill('solid', fgColor='FFD966'), font=Font(name='맑은 고딕', bold=True, size=12))

# ── 직접경비 ──
r = 15
ws2.cell(row=r, column=2, value='○ 직접경비').font = SEC_FONT
ws2.cell(row=r, column=12, value='(단위 : 원)').font = SMALL
r = 16
for c, h in enumerate(['구분', '', '', '산출내역', '', '', '', '', '', '', '금액'], 2):
    ws2.cell(row=r, column=c, value=h)
hdr(ws2, r, 12)
r = 17
ws2.cell(row=r, column=4, value='직접경비 요소 없음')
ws2.cell(row=r, column=11, value=0)
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, font=NF)
r = 18
ws2.cell(row=r, column=2, value='합 계').font = BF
ws2.cell(row=r, column=11, value=0)
ws2.cell(row=r, column=11).number_format = NUM
srow(ws2, r, 12, fill=SUB_FILL, font=BF)

# ── 보정계수 산정 ──
r = 20
ws2.cell(row=r, column=2, value='○ 보정계수 산정').font = SEC_FONT
for c, h in enumerate(['보정계수 구분', '', '', '복잡도 및 난이도 수준', '', '', '', '', '', '', '보정계수'], 2):
    ws2.cell(row=r+1, column=c, value=h) if c in [2,4,11] else None
# Headers
r = 20
ws2.cell(row=r, column=2, value='○ 보정계수 산정').font = SEC_FONT

corr_items = [
    (21, 'SW규모', "= 0.4057 x (log e(FP) - 7.1978)^2 + 0.8878\n   (단, 500FP 미만 1.28, 3,000FP 초과 0.866 적용)",
     "=IF('FP산정(간이법)'!H" + str(FP_TOTAL_ROW) + "<500,1.28,IF('FP산정(간이법)'!H" + str(FP_TOTAL_ROW) + ">3000,0.866,0.4057*(LN('FP산정(간이법)'!H" + str(FP_TOTAL_ROW) + ")-7.1978)^2+0.8878))"),
    (22, '연계복잡성', '4. 3~5개의 외부 시스템 연계\n   (원격DB + Ollama SLM + SNMP 13대 + Node-RED)', None),
    (23, '성능 요구수준', '3. 응답시간이나 처리율이 피크타임에 중요하며, 처리 시한이 명시\n   (실시간 모니터링 2초, SSE 스트리밍, 캐시 84→2초)', None),
    (24, '운영환경 호환성', '3. 유사 하드웨어 및 소프트웨어 환경에서 운영\n   (Docker + HTTPS + Cloudflare Tunnel)', None),
    (25, '보안성 요구수준', '2. 2가지 보안 요구사항 포함\n   (JWT 인증 + HTTPS + RBAC 역할 기반 접근제어)', None),
]

# 보정계수 값 (참고 문서 기준)
corr_values = {22: 1.00, 23: 1.00, 24: 1.06, 25: 1.00}

for row_n, label, desc, formula in corr_items:
    ws2.cell(row=row_n, column=2, value=label).font = BF
    ws2.cell(row=row_n, column=4, value=desc).alignment = L
    ws2.merge_cells(start_row=row_n, start_column=4, end_row=row_n, end_column=10)
    if formula:
        ws2.cell(row=row_n, column=11, value=formula)
    elif row_n in corr_values:
        ws2.cell(row=row_n, column=11, value=corr_values[row_n])
    ws2.cell(row=row_n, column=11).number_format = PCT
    ws2.cell(row=row_n, column=11).font = BLUE
    for c in range(2, 12):
        ws2.cell(row=row_n, column=c).border = BD
    ws2.row_dimensions[row_n].height = 35

# ── 보정계수 참조 테이블 (우측) ──
ref_data = [
    (22, [('1. 타 기관 연계 없음', 0.88), ('2. 1~2개의 타 기관 연계', 0.94),
          ('3. 3~5개의 타 기관 연계', 1.00), ('4. 6~10개의 타 기관 연계', 1.06),
          ('5. 10개 초과의 타 기관 연계', 1.12)]),
    (27, [('1. 응답성능에 대한 특별한 요구사항이 없다.', 0.91),
          ('2. 응답성능 요구사항이 있으나 특별한 조치 불필요', 0.95),
          ('3. 응답시간이 피크타임에 중요, 처리시한 명시', 1.00),
          ('4. 응답시간이 모든 업무시간에 중요, 처리시한 명시', 1.05),
          ('5. 성능 분석도구 사용이 필요할 정도로 엄격', 1.09)]),
    (33, [('1. 운영환경 호환성 요구사항 없음', 0.94),
          ('2. 동일 HW/SW 환경에서 운영', 1.00),
          ('3. 유사 HW/SW 환경에서 운영', 1.06),
          ('4. 이질적인 HW/SW 환경에서 운영', 1.13),
          ('5. 추가 운영 절차 문서화+모의훈련 요구', 1.19)]),
    (38, [('1. 1가지 보안 요구사항 포함', 0.97),
          ('2. 2가지 보안 요구사항 포함', 1.00),
          ('3. 3가지 보안 요구사항 포함', 1.03),
          ('4. 4가지 이상 보안 요구사항 포함', 1.06)]),
]

for start_row, items in ref_data:
    for i, (text, val) in enumerate(items):
        rr = start_row + i
        ws2.cell(row=rr, column=14, value=text).font = SMALL
        ws2.cell(row=rr, column=15, value=val).font = SMALL
        ws2.cell(row=rr, column=15).number_format = PCT

# ── FP 집계 ──
r = 27
ws2.cell(row=r-1, column=2, value='○ FP 집계').font = SEC_FONT
for c, h in enumerate(['구분', '', '', '', '  ILF', '  EIF', '  EI', '  EO', '  EQ', '  계'], 2):
    ws2.cell(row=r, column=c, value=h)
hdr(ws2, r, 11)

r = 28
ws2.cell(row=r, column=2, value='복잡도')
for c, (tp, w) in enumerate(FP_W.items(), 6):
    ws2.cell(row=r, column=c, value=w)
    ws2.cell(row=r, column=c).number_format = '0.0'
srow(ws2, r, 11, font=NF)

r = 29
ws2.cell(row=r, column=2, value='기능수')
fp_type_list = list(FP_W.keys())
for c, tp in enumerate(fp_type_list, 6):
    ws2.cell(row=r, column=c, value=fp_counts[tp])
ws2.cell(row=r, column=11, value=f'=SUM(F29:J29)')
srow(ws2, r, 11, font=BF)

r = 30
ws2.cell(row=r, column=2, value='기능점수')
for c, tp in enumerate(fp_type_list, 6):
    ws2.cell(row=r, column=c, value=f'=F28*F29' if c==6 else f'={get_column_letter(c)}28*{get_column_letter(c)}29')
    ws2.cell(row=r, column=c).number_format = '0.0'
ws2.cell(row=r, column=11, value='=SUM(F30:J30)')
ws2.cell(row=r, column=11).number_format = '0.0'
srow(ws2, r, 11, fill=SUB_FILL, font=BF)

r = 31
ws2.cell(row=r, column=2, value='비중')
for c in range(6, 11):
    ws2.cell(row=r, column=c, value=f'=IF(K30>0,{get_column_letter(c)}30/K30,0)')
    ws2.cell(row=r, column=c).number_format = '0.00%'
ws2.cell(row=r, column=11, value='=SUM(F31:J31)')
ws2.cell(row=r, column=11).number_format = '0.00%'
srow(ws2, r, 11, font=NF)

widths(ws2, [2, 16, 4, 12, 4, 8, 8, 8, 8, 8, 18, 4, 2, 45, 8])

# ═══════════════════════════════════════════════════════════════
# Sheet 3: DB 테이블 목록
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('DB 테이블 목록')
ws3.sheet_properties.tabColor = 'ED7D31'
title(ws3, 'DB 설계 목록 (PostgreSQL 16 + TimescaleDB)', mc=7)

headers3 = ['No', '카테고리', '테이블/뷰명', '유형', '설명', '주요 컬럼', '비고']
r = 3
for c, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=c, value=h)
hdr(ws3, r, 7)

tables = [
    ('인증', 'tb_auth', '테이블', '권한 그룹', 'auth_idn, auth_name, auth_level', ''),
    ('', 'tb_user', '테이블', '사용자 계정', 'user_id, pw, pw_migrated, auth_idn, region', 'bcrypt'),
    ('', 'tb_user_session', '테이블', '로그인 세션', 'session_id, user_id, expires_at', ''),
    ('', 'tb_access_log', '테이블', 'API 접근 로그', 'log_id, user_id, api_path, timestamp', ''),
    ('메뉴', 'tb_menu', '테이블', '메뉴 트리', 'menu_idn, pmenu_idn, menu_name, app_path', '자기참조'),
    ('', 'tb_auth_menu', '테이블', '권한↔메뉴', 'auth_idn, menu_idn, read_yn, write_yn', ''),
    ('', 'tb_menu_api', '테이블', '메뉴↔API', 'menu_idn, api_path, method', ''),
    ('코드', 'tb_grp_code', '테이블', '그룹 코드', 'grp_code, grp_name', ''),
    ('', 'tb_comm_code', '테이블', '상세 코드', 'grp_code, comm_code, comm_name', ''),
    ('채팅', 'tb_ai_chat_ask_group', '테이블', '채팅 세션', 'group_id, user_id, title', ''),
    ('', 'tb_ai_chat_ask', '테이블', '사용자 질문', 'ask_id, group_id, question, ask_at', ''),
    ('', 'tb_ai_chat_bot', '테이블', '봇 응답', 'bot_id, ask_id, answer, visual_data', 'JSONB'),
    ('', 'tb_ai_chat_ask_image', '테이블', '질문 이미지', 'image_id, ask_id, file_path', ''),
    ('', 'tb_ai_chat_bot_image', '테이블', '응답 이미지', 'image_id, bot_id, file_path', ''),
    ('', 'tb_ai_chat_faq', '테이블', 'FAQ', 'faq_id, question, category', ''),
    ('시설', 'tb_service_reservoir_info', '테이블', '배수지 정보', 'sitename, region, capacity_m3', 'JSONB'),
    ('', 'tb_service_reservoir_status', '테이블', '배수지 상태', 'sitename, hh, ll, target_level, meta', '16항목'),
    ('', 'tb_service_booster_station_info', '테이블', '가압장 정보', 'sitename, region, facilitytype', ''),
    ('', 'tb_service_booster_station_status', '테이블', '가압장 상태', 'sitename, hh, ll, meta', '26항목'),
    ('', 'tb_pressure_reducing_facility_info', '테이블', '감압시설 정보', 'sitename, region', ''),
    ('', 'tb_pressure_reducing_facility_status', '테이블', '감압시설 상태', 'sitename, hh, ll, meta', ''),
    ('', 'tb_block_info', '테이블', '블록 정보', 'sitename, block_level', ''),
    ('', 'tb_block_status', '테이블', '블록 상태', 'sitename, target_flow, meta', ''),
    ('설비', 'tb_equipment_info', '테이블', '설비 마스터', 'equipment_id, sitename, equipment_type, ip', '290건'),
    ('', 'tb_equipment_status', '테이블', '설비 상태', 'equipment_id, is_alive, rtt_ms', ''),
    ('', 'tb_equipment_alarm_report', '테이블', '경보 리포트', 'alarm_id, equipment_id, alarm_type', ''),
    ('', 'tb_equipment_tag_map', '테이블', '설비↔태그', 'equipment_id, tagsn, mapped_by', '3,375건'),
    ('네트워크', 'tb_network_link', '테이블', '네트워크 연결', 'source_id, target_id, protocol', ''),
    ('', 'tb_facility_flow_map', '테이블', '용수 계통도', 'source_sitename, target_sitename', '95엣지'),
    ('태그', 'tb_tag_info', '테이블', '태그 마스터', 'tagsn, sitename, datainfo, unit', '2,698건'),
    ('', 'tb_tag_raw_data', '하이퍼테이블', '시계열 데이터', 'tagsn, timestamp, value, quality', '2.3M+행'),
    ('', 'tb_tag_data_group', '테이블', '태그 그룹', 'group_code, parent_code, keywords', '21그룹'),
    ('', 'tb_tag_group_map', '테이블', '태그↔그룹', 'tagsn, group_code, matched_keyword', '93%자동'),
    ('', 'tb_trend_catalog', '테이블', '트렌드 카탈로그', 'catalog_id, name, items', ''),
    ('', 'tb_alarm_log', '테이블', '알람 이력', 'tagsn, alarm_type, start_time', ''),
    ('모니터링', 'tb_monitoring_catalog', '테이블', '모니터링 카탈로그', 'catalog_id, catalog_name, items', ''),
    ('', 'tb_admin_site_settings', '테이블', '사이트 설정', 'setting_key, setting_value', ''),
    ('', 'tb_site_anomaly_profile', '테이블', '현장 프로파일', 'sitename, group, p95, p05', '일1회'),
    ('캔버스', 'tb_canvas_node_position', '테이블', '노드 위치', 'sitename, facilitytype, x, y', ''),
    ('', 'tb_causal_chain_override', '테이블', '인과 오버라이드', 'sitename, facilitytype, zone, chain', 'JSONB'),
    ('', 'tb_snmp_port_status', '테이블', 'SNMP 포트', 'equipment_id, port_index, status', ''),
    ('관리', 'tb_prompt_template', '테이블', '프롬프트 템플릿', 'intent, template', ''),
    ('', 'tb_prompt_column', '테이블', '프롬프트 컬럼', 'intent, column_name', ''),
    ('', 'tb_file_storage', '테이블', '파일 메타데이터', 'file_id, file_path, mime_type', ''),
    ('', 'tb_file_history', '테이블', '파일 버전 이력', 'history_id, file_id, version', ''),
    ('', 'tb_facility_file', '테이블', '시설↔파일', 'facility_file_id, sitename, file_id', ''),
    ('', 'tb_field_lock', '테이블', '컬럼 잠금', 'table_name, column_name', ''),
    ('뷰/집계', 'v_reservoir_info_status', '뷰', '배수지 통합뷰', '', ''),
    ('', 'v_booster_station_info_status', '뷰', '가압장 통합뷰', '', ''),
    ('', 'v_pressure_reducing_facility_info_status', '뷰', '감압시설 통합뷰', '', ''),
    ('', 'v_block_info_status', '뷰', '블록 통합뷰', '', ''),
    ('', 'v_ongoing_alarm', '뷰', '진행 중 알람', '', ''),
    ('', 'cagg_5min_raw_stats_ai', '연속집계', '5분 통계', '', 'TimescaleDB'),
    ('', 'mv_tag_daily_status', '물리화뷰', '일별 통계', '', ''),
]

for i, t in enumerate(tables):
    r = 4 + i
    ws3.cell(row=r, column=1, value=i+1)
    for c, v in enumerate(t, 2):
        ws3.cell(row=r, column=c, value=v)
    type_cell = ws3.cell(row=r, column=4)
    if t[2] == '하이퍼테이블': type_cell.fill = PatternFill('solid', fgColor='E8D5F5')
    elif t[2] in ('뷰', '연속집계', '물리화뷰'): type_cell.fill = PatternFill('solid', fgColor='D5E8F5')
    als = [C, L, L, C, L, L, L]
    srow(ws3, r, 7, fill=CAT_FILL if t[0] else None, font=BF if t[0] else NF, als=als)

r_db = 4 + len(tables)
tbl_cnt = sum(1 for t in tables if t[2] == '테이블')
ws3.cell(row=r_db, column=1, value=f'합계: 테이블 {tbl_cnt}개 + 뷰/집계 {len(tables)-tbl_cnt}개 = {len(tables)}개')
ws3.merge_cells(start_row=r_db, start_column=1, end_row=r_db, end_column=7)
srow(ws3, r_db, 7, fill=TOTAL_FILL, font=BF)

widths(ws3, [5, 8, 38, 10, 20, 38, 12])

# ── 인쇄 설정 ──
for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

OUTPUT = 'D:/web/SLM_대가산정표_FP.xlsx'
wb.save(OUTPUT)

# 결과 요약
total_fp = sum(fp_counts[t] * FP_W[t] for t in fp_counts)
print(f'=== 저장 완료: {OUTPUT} ===')
print(f'\nFP 집계:')
for tp in fp_type_list:
    print(f'  {tp}: {fp_counts[tp]}개 × {FP_W[tp]} = {fp_counts[tp]*FP_W[tp]:.1f} FP')
print(f'  합계: {sum(fp_counts.values())}개 기능, {total_fp:.1f} FP')
print(f'\n보정 전 개발원가: {total_fp:.1f} × {FP_UNIT_PRICE:,} = {total_fp*FP_UNIT_PRICE:,.0f}원')
print(f'FP Total Row: {FP_TOTAL_ROW}')
