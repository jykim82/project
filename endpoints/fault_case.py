"""
tb_fault_case CRUD + 임베딩 + 엑셀 IMPORT/EXPORT 엔드포인트 [P3]

구조:
  - POST   /fault-cases           — 새 케이스 생성 + 임베딩
  - GET    /fault-cases           — 목록 (페이징 + equipment_type/brand 필터)
  - GET    /fault-cases/{id}      — 단건
  - PUT    /fault-cases/{id}      — 수정 + 임베딩 재생성
  - DELETE /fault-cases/{id}      — soft delete (is_active=false)
  - POST   /fault-cases/import    — 엑셀 업로드 (openpyxl)
  - GET    /fault-cases/export    — 엑셀 다운로드 (openpyxl)
  - GET    /fault-cases/template  — 빈 템플릿 다운로드

임베딩:
  - snowflake-arctic-embed2 (Ollama /api/embed) 1024차원, L2 normalized
  - 저장: data/fault_case_embeddings/fault_case_<id>.npz
  - vision_agent RAG 인덱스가 이 디렉터리를 로드해 검색 통합 (P3-F)

사양: docs/chat-photo-upload-scenario-spec.md §P3 / docs/work-history.md
"""

import io
import json
import logging
import os
from datetime import datetime
from typing import Any, Optional

import httpx
import numpy as np
from fastapi import APIRouter, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/fault-cases", tags=["fault-cases"])

# ─────────────────────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────────────────────

_get_db_connection = None

OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://host.docker.internal:11434")
EMBED_MODEL = os.environ.get("EMBED_MODEL", "snowflake-arctic-embed2:latest")
VISION_AGENT_URL = os.environ.get("VISION_AGENT_URL", "http://host.docker.internal:8100")

# 임베딩 NPZ 저장 경로 — vision_agent 도 동일 기본값을 사용 (P3-F)
FAULT_CASE_EMBEDDINGS_DIR = os.environ.get(
    "FAULT_CASE_EMBEDDINGS_DIR",
    os.path.normpath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "data", "fault_case_embeddings",
    )),
)

EQUIPMENT_TYPES = (
    "PLC", "유량계", "모뎀", "RTU", "인버터", "펌프", "밸브",
    "수위계", "압력계", "UPS", "기타",
)
SEVERITIES = ("경고", "주의", "정보")

# 엑셀 헤더 (IMPORT/EXPORT 공통 순서)
EXCEL_HEADERS = [
    "equipment_type", "brand", "model", "symptom",
    "cause", "action", "severity", "reference_url", "notes",
]


def init(get_db_connection_fn):
    global _get_db_connection
    _get_db_connection = get_db_connection_fn
    os.makedirs(FAULT_CASE_EMBEDDINGS_DIR, exist_ok=True)
    logger.info(f"fault_case init: embeddings_dir={FAULT_CASE_EMBEDDINGS_DIR}")


def _get_conn():
    if _get_db_connection is None:
        raise HTTPException(500, "DB 커넥션 미초기화")
    return _get_db_connection()


# ─────────────────────────────────────────────────────────────────────
# 임베딩 — snowflake-arctic-embed2
# ─────────────────────────────────────────────────────────────────────

def _build_embed_text(symptom: str, cause: Optional[str], action: Optional[str]) -> str:
    """증상 + 원인 + 조치 결합 — 임베딩 품질 향상을 위해 모두 합침."""
    parts = [symptom or ""]
    if cause:
        parts.append(cause)
    if action:
        parts.append(action)
    return " ".join(p.strip() for p in parts if p and p.strip())


def _embed_text(text: str) -> np.ndarray:
    """Ollama /api/embed → L2 normalized float32 (1024,)."""
    with httpx.Client(timeout=30.0) as client:
        resp = client.post(
            f"{OLLAMA_URL}/api/embed",
            json={"model": EMBED_MODEL, "input": text},
        )
        resp.raise_for_status()
        data = resp.json()
    emb = np.array(data["embeddings"][0], dtype=np.float32)
    norm = float(np.linalg.norm(emb))
    if norm > 0:
        emb = emb / norm
    return emb


def _save_embedding(case_id: int, embed_text: str) -> str:
    """임베딩 생성 + NPZ 저장. 반환: embedding_key(NPZ 파일명 prefix)."""
    emb = _embed_text(embed_text)
    key = f"fault_case_{case_id}"
    path = os.path.join(FAULT_CASE_EMBEDDINGS_DIR, f"{key}.npz")
    np.savez(path, embedding=emb, text=embed_text)
    return key


def _delete_embedding(embedding_key: Optional[str]) -> None:
    if not embedding_key:
        return
    path = os.path.join(FAULT_CASE_EMBEDDINGS_DIR, f"{embedding_key}.npz")
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError as e:
            logger.warning(f"임베딩 파일 삭제 실패: {path} {e}")


def _notify_vision_agent_reload() -> None:
    """CRUD/Import 직후 vision_agent 의 fault_case 인덱스 재로드 트리거.

    실패는 warning 으로만 처리 — 다음 진단 요청 시 자연 재로드 될 때까지
    지연될 수 있음.
    """
    try:
        with httpx.Client(timeout=3.0) as client:
            client.post(f"{VISION_AGENT_URL}/vision/fault-cases/reload")
    except Exception as e:
        logger.info(f"vision_agent reload skip: {e}")


# ─────────────────────────────────────────────────────────────────────
# Request/Response 모델
# ─────────────────────────────────────────────────────────────────────

class FaultCaseBase(BaseModel):
    equipment_type: str = Field(..., description="PLC/유량계/모뎀/...")
    brand: Optional[str] = None
    model: Optional[str] = None
    symptom: str
    cause: Optional[str] = None
    action: Optional[str] = None
    severity: Optional[str] = None
    reference_url: Optional[str] = None
    notes: Optional[str] = None


class FaultCaseCreate(FaultCaseBase):
    pass


class FaultCaseUpdate(BaseModel):
    equipment_type: Optional[str] = None
    brand: Optional[str] = None
    model: Optional[str] = None
    symptom: Optional[str] = None
    cause: Optional[str] = None
    action: Optional[str] = None
    severity: Optional[str] = None
    reference_url: Optional[str] = None
    notes: Optional[str] = None


class FaultCase(FaultCaseBase):
    case_id: int
    is_active: bool
    created_by: Optional[str] = None
    created_at: str
    updated_at: str
    embedding_updated_at: Optional[str] = None


class FaultCaseListResponse(BaseModel):
    items: list[FaultCase]
    total: int
    page: int
    page_size: int


def _validate_case(data: dict) -> None:
    et = (data.get("equipment_type") or "").strip()
    if et not in EQUIPMENT_TYPES:
        raise HTTPException(400, f"equipment_type 는 {EQUIPMENT_TYPES} 중 하나여야 합니다. 받음: {et!r}")
    sev = data.get("severity")
    if sev and sev not in SEVERITIES:
        raise HTTPException(400, f"severity 는 {SEVERITIES} 중 하나여야 합니다. 받음: {sev!r}")
    if not (data.get("symptom") or "").strip():
        raise HTTPException(400, "symptom 은 필수입니다")


def _row_to_case(row: tuple, cols: list[str]) -> FaultCase:
    d = dict(zip(cols, row))
    return FaultCase(
        case_id=d["case_id"],
        equipment_type=d["equipment_type"],
        brand=d.get("brand"),
        model=d.get("model"),
        symptom=d["symptom"],
        cause=d.get("cause"),
        action=d.get("action"),
        severity=d.get("severity"),
        reference_url=d.get("reference_url"),
        notes=d.get("notes"),
        is_active=bool(d.get("is_active")),
        created_by=d.get("created_by"),
        created_at=d["created_at"].isoformat() if d.get("created_at") else "",
        updated_at=d["updated_at"].isoformat() if d.get("updated_at") else "",
        embedding_updated_at=d["embedding_updated_at"].isoformat() if d.get("embedding_updated_at") else None,
    )


_BASE_SELECT = (
    "SELECT case_id, equipment_type, brand, model, symptom, cause, action, severity, "
    "reference_url, notes, is_active, created_by, created_at, updated_at, embedding_updated_at "
    "FROM tb_fault_case"
)
_BASE_COLS = ["case_id","equipment_type","brand","model","symptom","cause","action","severity",
              "reference_url","notes","is_active","created_by","created_at","updated_at","embedding_updated_at"]


# ─────────────────────────────────────────────────────────────────────
# CRUD 엔드포인트
# ─────────────────────────────────────────────────────────────────────

@router.get("", response_model=FaultCaseListResponse)
def list_fault_cases(
    equipment_type: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None, description="symptom/cause/action ILIKE"),
    include_inactive: bool = Query(False),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
):
    """목록 조회 (필터 + 페이징)."""
    conn = _get_conn()
    try:
        cur = conn.cursor()
        where = []
        params: list[Any] = []
        if not include_inactive:
            where.append("is_active = TRUE")
        if equipment_type:
            where.append("equipment_type = %s")
            params.append(equipment_type)
        if brand:
            where.append("brand ILIKE %s")
            params.append(f"%{brand}%")
        if keyword:
            where.append("(symptom ILIKE %s OR cause ILIKE %s OR action ILIKE %s)")
            params.extend([f"%{keyword}%"] * 3)
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        cur.execute(f"SELECT COUNT(*) FROM tb_fault_case{where_sql}", params)
        total = cur.fetchone()[0]
        offset = (page - 1) * page_size
        cur.execute(
            f"{_BASE_SELECT}{where_sql} ORDER BY case_id DESC LIMIT %s OFFSET %s",
            params + [page_size, offset],
        )
        rows = cur.fetchall()
        cur.close()
        items = [_row_to_case(r, _BASE_COLS) for r in rows]
        return FaultCaseListResponse(items=items, total=total, page=page, page_size=page_size)
    finally:
        conn.close()


@router.get("/template")
def export_template():
    """빈 엑셀 템플릿 — 헤더만 + 1행 예시."""
    wb = Workbook()
    ws = wb.active
    ws.title = "fault_cases"
    ws.append(EXCEL_HEADERS)
    # 예시 1행
    ws.append([
        "PLC", "LS", "XGB-XBCH",
        "ERR LED 점등 (빨강)",
        "CPU 펌웨어 이상 또는 전원 불안정",
        "전원 재기동 후 매뉴얼 6.3절 고장진단 절차 수행",
        "경고", "", "현장 최빈 케이스",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="fault_case_template.xlsx"'},
    )


@router.get("/export")
def export_fault_cases(
    equipment_type: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
):
    """현재 DB 케이스 → 엑셀 다운로드."""
    conn = _get_conn()
    try:
        cur = conn.cursor()
        where = []
        params: list[Any] = []
        if not include_inactive:
            where.append("is_active = TRUE")
        if equipment_type:
            where.append("equipment_type = %s")
            params.append(equipment_type)
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        cur.execute(f"{_BASE_SELECT}{where_sql} ORDER BY equipment_type, brand, model, case_id", params)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "fault_cases"
    ws.append(EXCEL_HEADERS)
    for r in rows:
        d = dict(zip(_BASE_COLS, r))
        ws.append([
            d["equipment_type"], d.get("brand") or "", d.get("model") or "",
            d["symptom"], d.get("cause") or "", d.get("action") or "",
            d.get("severity") or "", d.get("reference_url") or "", d.get("notes") or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"fault_cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/import")
async def import_fault_cases(
    file: UploadFile = File(...),
    user_id: Optional[str] = Query(None),
    overwrite: bool = Query(False, description="중복 symptom 덮어쓰기 (false면 skip)"),
):
    """엑셀 IMPORT — 헤더는 EXCEL_HEADERS 순서여야 함.

    결과: {"imported": N, "skipped": M, "errors": [{"row": i, "reason": ...}]}
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "xlsx 파일만 지원합니다")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 파일 로드 실패: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "빈 파일")
    header = [str(h or "").strip() for h in rows[0]]
    missing = [h for h in EXCEL_HEADERS if h not in header]
    if missing:
        raise HTTPException(400, f"누락된 헤더: {missing}")
    idx = {h: header.index(h) for h in EXCEL_HEADERS}

    imported = 0
    skipped = 0
    errors: list[dict] = []

    conn = _get_conn()
    try:
        cur = conn.cursor()
        for i, row in enumerate(rows[1:], start=2):  # 엑셀 행번호 (헤더 1행 제외)
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                data = {
                    h: (str(row[idx[h]]).strip() if row[idx[h]] is not None else None)
                    for h in EXCEL_HEADERS
                }
                # 빈 값 정리
                for k, v in list(data.items()):
                    if v == "":
                        data[k] = None
                _validate_case(data)
                # 중복 체크
                cur.execute(
                    "SELECT case_id FROM tb_fault_case "
                    "WHERE equipment_type=%s AND COALESCE(brand,'')=COALESCE(%s,'') "
                    "  AND COALESCE(model,'')=COALESCE(%s,'') AND symptom=%s AND is_active=TRUE",
                    (data["equipment_type"], data["brand"], data["model"], data["symptom"]),
                )
                existing = cur.fetchone()
                if existing and not overwrite:
                    skipped += 1
                    continue
                if existing and overwrite:
                    case_id = existing[0]
                    cur.execute(
                        "UPDATE tb_fault_case SET cause=%s, action=%s, severity=%s, "
                        "reference_url=%s, notes=%s WHERE case_id=%s",
                        (data["cause"], data["action"], data["severity"],
                         data["reference_url"], data["notes"], case_id),
                    )
                else:
                    cur.execute(
                        "INSERT INTO tb_fault_case "
                        "(equipment_type, brand, model, symptom, cause, action, severity, "
                        "reference_url, notes, created_by) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING case_id",
                        (data["equipment_type"], data["brand"], data["model"],
                         data["symptom"], data["cause"], data["action"],
                         data["severity"], data["reference_url"], data["notes"], user_id),
                    )
                    case_id = cur.fetchone()[0]
                # 임베딩 갱신
                try:
                    embed_text = _build_embed_text(data["symptom"], data["cause"], data["action"])
                    key = _save_embedding(case_id, embed_text)
                    cur.execute(
                        "UPDATE tb_fault_case SET embedding_key=%s, embedding_updated_at=NOW() WHERE case_id=%s",
                        (key, case_id),
                    )
                except Exception as e:
                    logger.warning(f"임베딩 생성 실패 case_id={case_id}: {e}")
                    # 임베딩 실패해도 DB 레코드는 유지 (나중에 re-embed 가능)
                imported += 1
            except HTTPException as e:
                errors.append({"row": i, "reason": e.detail})
            except Exception as e:
                errors.append({"row": i, "reason": str(e)[:200]})
        conn.commit()
        cur.close()
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"import 실패: {e}")
    finally:
        conn.close()

    if imported > 0:
        _notify_vision_agent_reload()
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows) - 1,
    }


@router.post("", response_model=FaultCase, status_code=201)
def create_fault_case(req: FaultCaseCreate, user_id: Optional[str] = Query(None)):
    data = req.model_dump()
    _validate_case(data)
    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO tb_fault_case "
            f"(equipment_type, brand, model, symptom, cause, action, severity, "
            f"reference_url, notes, created_by) "
            f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING case_id",
            (data["equipment_type"], data["brand"], data["model"],
             data["symptom"], data["cause"], data["action"],
             data["severity"], data["reference_url"], data["notes"], user_id),
        )
        case_id = cur.fetchone()[0]
        try:
            embed_text = _build_embed_text(data["symptom"], data["cause"], data["action"])
            key = _save_embedding(case_id, embed_text)
            cur.execute(
                "UPDATE tb_fault_case SET embedding_key=%s, embedding_updated_at=NOW() WHERE case_id=%s",
                (key, case_id),
            )
        except Exception as e:
            logger.warning(f"임베딩 생성 실패 case_id={case_id}: {e}")
        cur.execute(f"{_BASE_SELECT} WHERE case_id=%s", (case_id,))
        row = cur.fetchone()
        conn.commit()
        cur.close()
        _notify_vision_agent_reload()
        return _row_to_case(row, _BASE_COLS)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"create 실패: {e}")
    finally:
        conn.close()


@router.get("/{case_id}", response_model=FaultCase)
def get_fault_case(case_id: int):
    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"{_BASE_SELECT} WHERE case_id=%s", (case_id,))
        row = cur.fetchone()
        cur.close()
        if not row:
            raise HTTPException(404, "case_id not found")
        return _row_to_case(row, _BASE_COLS)
    finally:
        conn.close()


@router.put("/{case_id}", response_model=FaultCase)
def update_fault_case(case_id: int, req: FaultCaseUpdate):
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "변경할 필드가 없습니다")
    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"{_BASE_SELECT} WHERE case_id=%s", (case_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "case_id not found")
        current = dict(zip(_BASE_COLS, row))
        merged = {**current, **updates}
        _validate_case(merged)
        set_clauses = [f"{k}=%s" for k in updates.keys()]
        cur.execute(
            f"UPDATE tb_fault_case SET {', '.join(set_clauses)} WHERE case_id=%s",
            list(updates.values()) + [case_id],
        )
        # 임베딩 대상 필드가 바뀌었으면 재생성
        embed_affected = bool({"symptom","cause","action"} & set(updates.keys()))
        if embed_affected:
            try:
                embed_text = _build_embed_text(
                    merged.get("symptom") or "",
                    merged.get("cause"),
                    merged.get("action"),
                )
                key = _save_embedding(case_id, embed_text)
                cur.execute(
                    "UPDATE tb_fault_case SET embedding_key=%s, embedding_updated_at=NOW() WHERE case_id=%s",
                    (key, case_id),
                )
            except Exception as e:
                logger.warning(f"임베딩 재생성 실패 case_id={case_id}: {e}")
        cur.execute(f"{_BASE_SELECT} WHERE case_id=%s", (case_id,))
        updated_row = cur.fetchone()
        conn.commit()
        cur.close()
        _notify_vision_agent_reload()
        return _row_to_case(updated_row, _BASE_COLS)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"update 실패: {e}")
    finally:
        conn.close()


@router.delete("/{case_id}")
def delete_fault_case(case_id: int, hard: bool = Query(False)):
    """기본은 soft delete (is_active=FALSE). hard=True 면 완전 삭제."""
    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT embedding_key FROM tb_fault_case WHERE case_id=%s", (case_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "case_id not found")
        embedding_key = row[0]
        if hard:
            cur.execute("DELETE FROM tb_fault_case WHERE case_id=%s", (case_id,))
            _delete_embedding(embedding_key)
        else:
            cur.execute("UPDATE tb_fault_case SET is_active=FALSE WHERE case_id=%s", (case_id,))
        conn.commit()
        cur.close()
        _notify_vision_agent_reload()
        return {"status": "deleted" if hard else "deactivated", "case_id": case_id}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"delete 실패: {e}")
    finally:
        conn.close()
